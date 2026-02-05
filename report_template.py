"""
HTML Report Generation Template for PID Annotator
Generates interactive HTML reports with search, sort, filter, and CSV/Excel export functionality

Usage Examples:
--------------

1. Generate HTML Report (includes client-side Excel export via JavaScript):
    from report_template import generate_html_report

    html_content = generate_html_report(
        report_data={
            'found': [...],
            'not_found': [...],
            'duplicates': {...},
            'validation_warnings': [...]
        },
        pdf_filenames=['drawing1.pdf', 'drawing2.pdf'],
        excel_filename='component_list.xlsx',
        settings={'tag_column': 'A', 'header_row': 6}
    )

2. Server-side Excel Export (for Flask routes):
    from report_template import export_filtered_list_to_excel
    from flask import send_file

    @app.route('/export_found_tags')
    def export_found():
        found_tags = session.get('found_tags', [])
        excel_buffer = export_filtered_list_to_excel(found_tags, 'found')
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='found_tags.xlsx'
        )
"""

import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO


def export_filtered_list_to_excel(data_list, list_type='found', include_metadata=True):
    """
    Export filtered tag list to Excel format with proper formatting and headers.

    Args:
        data_list: List of tag data (found, not_found, duplicates, etc.)
        list_type: Type of list ('found', 'not_found', 'duplicates', 'validation_warnings')
        include_metadata: Whether to include metadata header rows (default: True)

    Returns:
        BytesIO: Excel file in memory as bytes
    """
    wb = Workbook()
    ws = wb.active

    # Set worksheet title based on list type
    title_map = {
        'found': 'Found Tags',
        'not_found': 'Not Found Tags',
        'duplicates': 'Duplicate Tags',
        'validation_warnings': 'Validation Warnings'
    }
    ws.title = title_map.get(list_type, 'Tag List')

    # Define styles
    header_fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    metadata_fill = PatternFill(start_color='f3f4f6', end_color='f3f4f6', fill_type='solid')
    metadata_font = Font(bold=True, size=10)
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row_idx = 1

    # Add metadata if requested
    if include_metadata:
        ws.merge_cells(f'A{row_idx}:D{row_idx}')
        metadata_cell = ws[f'A{row_idx}']
        metadata_cell.value = f'PID Annotator - {title_map.get(list_type, "Export")}'
        metadata_cell.font = Font(bold=True, size=14, color='2563eb')
        metadata_cell.alignment = Alignment(horizontal='center', vertical='center')
        row_idx += 1

        ws.merge_cells(f'A{row_idx}:D{row_idx}')
        date_cell = ws[f'A{row_idx}']
        date_cell.value = f'Exported: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        date_cell.font = Font(size=10, color='6b7280')
        date_cell.alignment = Alignment(horizontal='center')
        row_idx += 2  # Add blank row

    # Define headers based on list type
    if list_type == 'found':
        headers = ['Tag', 'Pages', 'Bookmarks', 'Occurrences', 'Excel Row']
    elif list_type == 'not_found':
        headers = ['Tag', 'Excel Row', 'Reason']
    elif list_type == 'duplicates':
        headers = ['Tag', 'Excel Rows', 'Count']
    elif list_type == 'validation_warnings':
        headers = ['Tag', 'Excel Row', 'Warning']
    else:
        headers = ['Tag', 'Data']

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin

    header_row = row_idx
    row_idx += 1

    # Write data rows
    for item in data_list:
        if list_type == 'found':
            # Get bookmarks and format them
            bookmarks = item.get('bookmarks', [])
            bookmarks_str = ', '.join(str(b) for b in bookmarks) if bookmarks else 'N/A'

            row_data = [
                item.get('tag', ''),
                ', '.join(map(str, item.get('pages', []))),
                bookmarks_str,
                item.get('occurrence_count', 0),
                item.get('excel_row', '')
            ]
        elif list_type == 'not_found':
            row_data = [
                item.get('tag', ''),
                item.get('excel_row', ''),
                item.get('reason', 'not_in_pdf')
            ]
        elif list_type == 'duplicates':
            row_data = [
                item.get('tag', ''),
                ', '.join(map(str, item.get('excel_rows', []))),
                item.get('count', 0)
            ]
        elif list_type == 'validation_warnings':
            row_data = [
                item.get('tag', ''),
                item.get('excel_row', ''),
                item.get('warning', '')
            ]
        else:
            row_data = [str(item)]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border_thin
            cell.alignment = Alignment(vertical='center')

            # Special formatting for tag column
            if col_idx == 1:
                cell.font = Font(name='Courier New', bold=True, color='2563eb')

        row_idx += 1

    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, 1):
        max_length = len(header)
        column_letter = ws.cell(row=header_row, column=col_idx).column_letter

        # Check data rows for maximum width
        for row in ws.iter_rows(min_row=header_row + 1, max_row=row_idx - 1, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

        # Set column width with padding (max 50 characters)
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer


def generate_html_report(report_data, pdf_filenames, excel_filename, settings=None):
    """
    Generate interactive HTML report with search, sort, filter, and CSV export.

    Args:
        report_data: Dict with 'found', 'not_found', 'duplicates', 'validation_warnings'
        pdf_filenames: List of PDF file names processed
        excel_filename: Excel file name used
        settings: Dict of processing settings (optional)

    Returns:
        str: Complete HTML document as string
    """
    # Prepare data
    found = report_data.get('found', [])
    not_found = report_data.get('not_found', [])
    duplicates = report_data.get('duplicates', {})
    validation_warnings = report_data.get('validation_warnings', [])
    page_stats = report_data.get('page_stats', {})
    unmatched_pdf_tags = report_data.get('unmatched_pdf_tags', [])

    total_tags = len(found) + len(not_found)
    found_count = len(found)
    not_found_count = len(not_found)
    duplicate_count = len(duplicates)
    warning_count = len(validation_warnings)
    unmatched_count = len(unmatched_pdf_tags)

    found_percent = (found_count / total_tags * 100) if total_tags > 0 else 0
    not_found_percent = (not_found_count / total_tags * 100) if total_tags > 0 else 0

    # Current timestamp
    process_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # PDF file list
    pdf_list = ', '.join(pdf_filenames) if isinstance(pdf_filenames, list) else str(pdf_filenames)

    # Convert duplicates dict to list for easier iteration
    duplicates_list = []
    for tag, data in duplicates.items():
        # Handle both old format (list of rows) and new format (dict with rows and data)
        if isinstance(data, dict):
            excel_rows = data.get('excel_rows', [])
            row_data = data.get('row_data', [])
        else:
            # Legacy format - just a list of row numbers
            excel_rows = data
            row_data = []

        duplicates_list.append({
            'tag': tag,
            'excel_rows': excel_rows,
            'count': len(excel_rows),
            'row_data': row_data
        })

    # Build found rows with dropdown panels
    found_rows = []
    for idx, item in enumerate(found):
        tag = item["tag"]
        pages = [p+1 for p in item["pages"]]
        bookmarks = item.get("bookmarks", ["N/A"])
        occurrence_count = item["occurrence_count"]
        excel_row = item["excel_row"]
        row_data = item.get("row_data", {})

        detail_id = f"foundDetail{idx}"
        has_details = len(row_data) > 0

        # Build toggle button if we have row data
        toggle_button_html = ''
        if has_details:
            toggle_button_html = (
                f'<button type="button" class="found-detail-toggle" '
                f'aria-expanded="false" aria-controls="{detail_id}" '
                f'onclick="return toggleFoundDetail(\'{detail_id}\', this);">'
                f'<span class="chevron">▼</span>'
                f'</button>'
            )

        # Main row
        row_attributes = 'class="found-summary-row"'
        if has_details:
            row_attributes += f' data-detail-id="{detail_id}"'

        row_html = f'''                    <tr {row_attributes}>
                        <td class="tag-name">
                            <div class="found-summary-cell">
                                {toggle_button_html}
                                <span>{tag}</span>
                            </div>
                        </td>
                        <td class="page-list">{", ".join(map(str, pages))}</td>
                        <td class="bookmark-list">{", ".join(str(b) for b in bookmarks)}</td>
                        <td>{occurrence_count}</td>
                        <td class="excel-row">{excel_row}</td>
                    </tr>'''

        # Detail panel row
        if has_details and row_data:
            # Get column headers from row data (excluding excel_row as it's shown separately)
            headers = [col for col in row_data.keys() if col != 'excel_row']

            # Build detail table with single row of data
            detail_table_cells = []
            detail_table_cells.append(f'<td class="excel-row">{row_data.get("excel_row", excel_row)}</td>')
            for header in headers:
                value = row_data.get(header, "")
                # Convert to string and escape HTML
                value_str = str(value) if value != "" else ""
                detail_table_cells.append(f'<td>{value_str}</td>')

            detail_row_html = f'''                                    <tr>
                                        {chr(10).join(detail_table_cells)}
                                    </tr>'''

            # Build header row for detail table
            header_cells = ['<th>Excel Row</th>'] + [f'<th>{h}</th>' for h in headers]
            header_row_html = '\n                                            '.join(header_cells)

            detail_row = f'''                    <tr id="{detail_id}" class="found-detail-row hidden">
                        <td colspan="5">
                            <div class="found-detail-panel">
                                <div class="found-detail-header">
                                    <span class="found-detail-title">Excel Data for Found Tag: {tag}</span>
                                    <span class="badge">{occurrence_count} occurrence(s) in PDF</span>
                                </div>
                                <table class="found-data-table">
                                    <thead>
                                        <tr>
                                            {header_row_html}
                                        </tr>
                                    </thead>
                                    <tbody>
{detail_row_html}
                                    </tbody>
                                </table>
                            </div>
                        </td>
                    </tr>'''
        else:
            detail_row = ''

        found_rows.append(row_html + ('\n' + detail_row if detail_row else ''))

    found_rows_html = '\n'.join(found_rows)

    # Build not_found rows with dropdown panels
    not_found_rows = []
    for idx, item in enumerate(not_found):
        tag = item["tag"]
        excel_row = item["excel_row"]
        reason = item.get("reason", "not_in_pdf")
        row_data = item.get("row_data", {})

        detail_id = f"notFoundDetail{idx}"
        has_details = len(row_data) > 0

        # Build toggle button if we have row data
        toggle_button_html = ''
        if has_details:
            toggle_button_html = (
                f'<button type="button" class="not-found-detail-toggle" '
                f'aria-expanded="false" aria-controls="{detail_id}" '
                f'onclick="return toggleNotFoundDetail(\'{detail_id}\', this);">'
                f'<span class="chevron">▼</span>'
                f'</button>'
            )

        # Main row
        row_attributes = 'class="not-found-summary-row"'
        if has_details:
            row_attributes += f' data-detail-id="{detail_id}"'

        row_html = f'''                    <tr {row_attributes}>
                        <td class="tag-name">
                            <div class="not-found-summary-cell">
                                {toggle_button_html}
                                <span>{tag}</span>
                            </div>
                        </td>
                        <td class="excel-row">{excel_row}</td>
                        <td>{reason}</td>
                    </tr>'''

        # Detail panel row
        if has_details and row_data:
            # Get column headers from row data (excluding excel_row as it's shown separately)
            headers = [col for col in row_data.keys() if col != 'excel_row']

            # Build detail table with single row of data
            detail_table_cells = []
            detail_table_cells.append(f'<td class="excel-row">{row_data.get("excel_row", excel_row)}</td>')
            for header in headers:
                value = row_data.get(header, "")
                # Convert to string and escape HTML
                value_str = str(value) if value != "" else ""
                detail_table_cells.append(f'<td>{value_str}</td>')

            detail_row_html = f'''                                    <tr>
                                        {chr(10).join(detail_table_cells)}
                                    </tr>'''

            # Build header row for detail table
            header_cells = ['<th>Excel Row</th>'] + [f'<th>{h}</th>' for h in headers]
            header_row_html = '\n                                            '.join(header_cells)

            detail_row = f'''                    <tr id="{detail_id}" class="not-found-detail-row hidden">
                        <td colspan="3">
                            <div class="not-found-detail-panel">
                                <div class="not-found-detail-header">
                                    <span class="not-found-detail-title">Excel Data for Not Found Tag: {tag}</span>
                                    <span class="badge">{reason}</span>
                                </div>
                                <table class="not-found-data-table">
                                    <thead>
                                        <tr>
                                            {header_row_html}
                                        </tr>
                                    </thead>
                                    <tbody>
{detail_row_html}
                                    </tbody>
                                </table>
                            </div>
                        </td>
                    </tr>'''
        else:
            detail_row = ''

        not_found_rows.append(row_html + ('\n' + detail_row if detail_row else ''))

    not_found_rows_html = '\n'.join(not_found_rows)

    # Build duplicates rows with dropdown panels
    duplicates_rows = []
    for idx, item in enumerate(duplicates_list):
        tag = item["tag"]
        excel_rows = item["excel_rows"]
        count = item["count"]
        row_data = item.get("row_data", [])

        detail_id = f"duplicateDetail{idx}"
        has_details = len(row_data) > 0

        # Build toggle button if we have row data
        toggle_button_html = ''
        if has_details:
            toggle_button_html = (
                f'<button type="button" class="duplicate-detail-toggle" '
                f'aria-expanded="false" aria-controls="{detail_id}" '
                f'onclick="return toggleDuplicateDetail(\'{detail_id}\', this);">'
                f'<span class="chevron">▼</span>'
                f'</button>'
            )

        # Main row
        row_attributes = 'class="duplicate-summary-row"'
        if has_details:
            row_attributes += f' data-detail-id="{detail_id}"'

        row_html = f'''                    <tr {row_attributes}>
                        <td class="tag-name">
                            <div class="duplicate-summary-cell">
                                {toggle_button_html}
                                <span>{tag}</span>
                            </div>
                        </td>
                        <td class="excel-row">{", ".join(map(str, excel_rows))}</td>
                        <td>{count}</td>
                    </tr>'''

        # Detail panel row
        if has_details and row_data:
            # Get column headers from first row
            headers = [col for col in row_data[0].keys() if col != 'excel_row']

            # Build detail table rows
            detail_table_rows = []
            for row in row_data:
                cells = []
                cells.append(f'<td class="excel-row">{row.get("excel_row", "")}</td>')
                for header in headers:
                    value = row.get(header, "")
                    # Convert to string and escape HTML
                    value_str = str(value) if value != "" else ""
                    cells.append(f'<td>{value_str}</td>')
                detail_table_rows.append(f'''                                    <tr>
                                        {chr(10).join(cells)}
                                    </tr>''')

            detail_rows_html = '\n'.join(detail_table_rows)

            # Build header row for detail table
            header_cells = ['<th>Excel Row</th>'] + [f'<th>{h}</th>' for h in headers]
            header_row_html = '\n                                            '.join(header_cells)

            detail_row = f'''                    <tr id="{detail_id}" class="duplicate-detail-row hidden">
                        <td colspan="3">
                            <div class="duplicate-detail-panel">
                                <div class="duplicate-detail-header">
                                    <span class="duplicate-detail-title">Excel Data for Duplicated Tag: {tag}</span>
                                    <span class="badge">{count} occurrences</span>
                                </div>
                                <table class="duplicate-data-table">
                                    <thead>
                                        <tr>
                                            {header_row_html}
                                        </tr>
                                    </thead>
                                    <tbody>
{detail_rows_html}
                                    </tbody>
                                </table>
                            </div>
                        </td>
                    </tr>'''
        else:
            detail_row = ''

        duplicates_rows.append(row_html + ('\n' + detail_row if detail_row else ''))

    duplicates_rows_html = '\n'.join(duplicates_rows)

    # Build validation warnings rows with dropdown panels
    warnings_rows = []
    for idx, item in enumerate(validation_warnings):
        tag = item["tag"]
        excel_row = item["excel_row"]
        warning = item["warning"]
        row_data = item.get("row_data", {})

        detail_id = f"warningDetail{idx}"
        has_details = len(row_data) > 0

        # Build toggle button if we have row data
        toggle_button_html = ''
        if has_details:
            toggle_button_html = (
                f'<button type="button" class="warning-detail-toggle" '
                f'aria-expanded="false" aria-controls="{detail_id}" '
                f'onclick="return toggleWarningDetail(\'{detail_id}\', this);">'
                f'<span class="chevron">▼</span>'
                f'</button>'
            )

        # Main row
        row_attributes = 'class="warning-summary-row"'
        if has_details:
            row_attributes += f' data-detail-id="{detail_id}"'

        row_html = f'''                    <tr {row_attributes}>
                        <td class="tag-name">
                            <div class="warning-summary-cell">
                                {toggle_button_html}
                                <span>{tag}</span>
                            </div>
                        </td>
                        <td class="excel-row">{excel_row}</td>
                        <td class="warning-text">{warning}</td>
                    </tr>'''

        # Detail panel row
        if has_details and row_data:
            # Get column headers from row data (excluding excel_row as it's shown separately)
            headers = [col for col in row_data.keys() if col != 'excel_row']

            # Build detail table with single row of data
            detail_table_cells = []
            detail_table_cells.append(f'<td class="excel-row">{row_data.get("excel_row", excel_row)}</td>')
            for header in headers:
                value = row_data.get(header, "")
                # Convert to string and escape HTML
                value_str = str(value) if value != "" else ""
                detail_table_cells.append(f'<td>{value_str}</td>')

            detail_row_html = f'''                                    <tr>
                                        {chr(10).join(detail_table_cells)}
                                    </tr>'''

            # Build header row for detail table
            header_cells = ['<th>Excel Row</th>'] + [f'<th>{h}</th>' for h in headers]
            header_row_html = '\n                                            '.join(header_cells)

            detail_row = f'''                    <tr id="{detail_id}" class="warning-detail-row hidden">
                        <td colspan="3">
                            <div class="warning-detail-panel">
                                <div class="warning-detail-header">
                                    <span class="warning-detail-title">Excel Data for Warning Tag: {tag}</span>
                                    <span class="badge">{warning}</span>
                                </div>
                                <table class="warning-data-table">
                                    <thead>
                                        <tr>
                                            {header_row_html}
                                        </tr>
                                    </thead>
                                    <tbody>
{detail_row_html}
                                    </tbody>
                                </table>
                            </div>
                        </td>
                    </tr>'''
        else:
            detail_row = ''

        warnings_rows.append(row_html + ('\n' + detail_row if detail_row else ''))

    warnings_rows_html = '\n'.join(warnings_rows)

    # Build unmatched PDF tags rows (tags in PDF but not in Excel)
    unmatched_rows = []
    for idx, item in enumerate(unmatched_pdf_tags):
        tag = item["tag"]
        # Convert 0-indexed pages to 1-indexed for display
        pages = [p + 1 for p in item["pages"]]
        occurrence_count = item["occurrence_count"]

        pages_str = ", ".join(map(str, pages))

        row_html = f'''                    <tr>
                        <td class="tag-name">{tag}</td>
                        <td class="page-list">{pages_str}</td>
                        <td class="occurrence-count">{occurrence_count}</td>
                    </tr>'''

        unmatched_rows.append(row_html)

    unmatched_rows_html = '\n'.join(unmatched_rows)

    # Process page statistics
    def _flatten_page_stats_map(stats_map, source_label=None):
        entries = []
        if not isinstance(stats_map, dict):
            return entries

        def sort_key(key):
            try:
                return (0, int(key))
            except (TypeError, ValueError):
                return (1, str(key))

        for page_key in sorted(stats_map.keys(), key=sort_key):
            stats = stats_map.get(page_key, {})
            if not isinstance(stats, dict):
                continue

            try:
                page_index = int(page_key)
                page_display = page_index + 1
            except (TypeError, ValueError):
                page_index = None
                page_display = str(page_key)

            colored_count = stats.get('colored_count', 0)
            total_count = stats.get('total_count', 0)
            bookmark = stats.get('bookmark', 'N/A')
            percentage = (colored_count / total_count * 100) if total_count > 0 else 0

            entry = {
                'page': page_display,
                'page_index': page_index,
                'colored_count': colored_count,
                'total_count': total_count,
                'bookmark': bookmark,
                'percentage': percentage
            }

            if source_label:
                entry['source_pdf'] = source_label

            tag_details = stats.get('tag_details')
            if isinstance(tag_details, list):
                entry['tag_details'] = list(tag_details)

            entries.append(entry)

        return entries

    def _format_page_label(item, include_source=False):
        page_value = item.get('page')
        label = str(page_value)
        try:
            # Handle numeric strings gracefully (e.g., "3")
            page_num = int(page_value)
            label = f"Page {page_num}"
        except (TypeError, ValueError):
            label = str(page_value)

        if include_source and item.get('source_pdf'):
            label = f"{item['source_pdf']} — {label}"

        return label

    page_stats_list = []
    if isinstance(page_stats, dict) and page_stats:
        sample_value = next(iter(page_stats.values()))
        # Determine if this is a simple {page_num: stats} mapping or nested per-PDF structure
        if isinstance(sample_value, dict) and {'total_count', 'colored_count'}.issubset(sample_value.keys()):
            page_stats_list.extend(_flatten_page_stats_map(page_stats))
        else:
            for source_label, stats_map in page_stats.items():
                page_stats_list.extend(_flatten_page_stats_map(stats_map, source_label=str(source_label)))
    elif isinstance(page_stats, list):
        for item in page_stats:
            if isinstance(item, dict) and 'total_count' in item:
                entry = {
                    'page': item.get('page'),
                    'page_index': item.get('page_index'),
                    'colored_count': item.get('colored_count', 0),
                    'total_count': item.get('total_count', 0),
                    'bookmark': item.get('bookmark', 'N/A'),
                    'percentage': item.get('percentage', 0),
                }
                if 'source_pdf' in item:
                    entry['source_pdf'] = item['source_pdf']
                if isinstance(item.get('tag_details'), list):
                    entry['tag_details'] = list(item['tag_details'])
                page_stats_list.append(entry)

    include_source_pdf = any(item.get('source_pdf') for item in page_stats_list)

    page_stats_count = len(page_stats_list)

    # Calculate page stats summary
    if page_stats_list:
        total_pages_with_tags = len([p for p in page_stats_list if p['total_count'] > 0])
        total_pages_with_colored = len([p for p in page_stats_list if p['colored_count'] > 0])
        pages_no_colored = total_pages_with_tags - total_pages_with_colored
        most_tagged_page = max(page_stats_list, key=lambda x: x['total_count']) if page_stats_list else None
        avg_tags_per_page = sum(p['total_count'] for p in page_stats_list) / len(page_stats_list) if page_stats_list else 0
    else:
        total_pages_with_tags = 0
        total_pages_with_colored = 0
        pages_no_colored = 0
        most_tagged_page = None
        avg_tags_per_page = 0

    page_stats_rows = []
    page_stats_column_count = 5 + (1 if include_source_pdf else 0)
    for idx, item in enumerate(page_stats_list):
        source_column = ''
        if include_source_pdf:
            source_value = item.get('source_pdf', 'N/A')
            source_column = f"\n                        <td class=\"pdf-name\">{source_value}</td>"

        detail_id = f"pageStatsDetail{idx}"
        tag_details = item.get('tag_details') or []
        detail_count = len(tag_details)
        has_details = detail_count > 0
        detail_label = 'match' if detail_count == 1 else 'matches'

        toggle_button_html = ''
        if has_details:
            toggle_button_html = (
                f"<button type=\"button\" class=\"page-detail-toggle\" "
                f"aria-expanded=\"false\" aria-controls=\"{detail_id}\" "
                f"onclick=\"return togglePageDetail('{detail_id}', this);\">"
                f"<span class=\"chevron\">▼</span>"
                f"</button>"
            )

        row_attributes = 'class="page-summary-row"'
        if has_details:
            row_attributes += f' data-detail-id="{detail_id}"'

        row_html = f'''                    <tr {row_attributes}>
                        <td class="page-number">
                            <div class="page-summary-cell">
                                {toggle_button_html}
                                <span>{item["page"]}</span>
                            </div>
                        </td>{source_column}
                        <td class="bookmark-title">{item["bookmark"]}</td>
                        <td>{item["colored_count"]}</td>
                        <td>{item["total_count"]}</td>
                        <td>
                            <div class="percentage-row">
                                <div class="percentage-bar-container">
                                    <div class="percentage-bar" style="width: {item["percentage"]}%"></div>
                                </div>
                                <span class="percentage-value">{item["percentage"]:.1f}%</span>
                            </div>
                        </td>
                    </tr>'''

        if has_details:
            detail_entries = []
            for detail_idx, detail in enumerate(tag_details):
                detected_text = detail.get('found_text') or detail.get('tag') or 'N/A'
                normalized_tag = detail.get('tag') or 'N/A'
                excel_match = bool(detail.get('in_excel'))
                colored_match = bool(detail.get('colored'))
                excel_class = 'status-success' if excel_match else 'status-muted'
                excel_label = 'Matched' if excel_match else 'Not Matched'
                colored_class = 'status-success' if colored_match else 'status-warning'
                colored_label = 'Colored' if colored_match else 'Not Colored'

                # Extract coloring reason from the detail data
                coloring_reason = detail.get('coloring_reason', 'N/A')

                # Determine the CSS class for the coloring reason
                reason_class = 'reason-default'
                if coloring_reason and coloring_reason != 'N/A':
                    reason_lower = str(coloring_reason).lower()
                    if 'user' in reason_lower or 'manual' in reason_lower:
                        reason_class = 'reason-user'
                    elif 'auto' in reason_lower or 'matched' in reason_lower:
                        reason_class = 'reason-auto'
                    elif 'custom' in reason_lower or 'rule' in reason_lower:
                        reason_class = 'reason-custom'
                    elif 'not' in reason_lower or 'none' in reason_lower:
                        reason_class = 'reason-none'

                # Check if this tag has Excel row data
                row_data = detail.get('row_data', {})
                has_row_data = len(row_data) > 0
                tag_excel_detail_id = f"pageStatsTagDetail{idx}_{detail_idx}"

                # Build toggle button if we have row data
                toggle_button_html = ''
                if has_row_data:
                    toggle_button_html = (
                        f'<button type="button" class="page-tag-excel-toggle" '
                        f'aria-expanded="false" aria-controls="{tag_excel_detail_id}" '
                        f'onclick="return togglePageTagExcel(\'{tag_excel_detail_id}\', this);">'
                        f'<span class="chevron-small">▼</span>'
                        f'</button>'
                    )

                # Main tag row
                tag_row_html = f'''                                    <tr class="page-tag-summary-row" {('data-excel-detail-id="' + tag_excel_detail_id + '"') if has_row_data else ''}>
                                        <td>
                                            <div class="page-tag-cell">
                                                {toggle_button_html}
                                                <span>{detected_text}</span>
                                            </div>
                                        </td>
                                        <td class="detail-tag">{normalized_tag}</td>
                                        <td><span class="status-badge {excel_class}">{excel_label}</span></td>
                                        <td><span class="status-badge {colored_class}">{colored_label}</span></td>
                                        <td><span class="coloring-reason {reason_class}">{coloring_reason}</span></td>
                                    </tr>'''

                detail_entries.append(tag_row_html)

                # Add Excel row data detail row if available
                if has_row_data:
                    # Get column headers from row data (excluding excel_row as it's shown separately)
                    headers = [col for col in row_data.keys() if col != 'excel_row']

                    # Build detail table with single row of data
                    detail_table_cells = []
                    detail_table_cells.append(f'<td class="excel-row">{row_data.get("excel_row", "")}</td>')
                    for header in headers:
                        value = row_data.get(header, "")
                        # Convert to string and escape HTML
                        value_str = str(value) if value != "" else ""
                        detail_table_cells.append(f'<td>{value_str}</td>')

                    detail_row_html = f'''                                    <tr>
                                        {chr(10).join(detail_table_cells)}
                                    </tr>'''

                    # Build header row for detail table
                    header_cells = ['<th>Excel Row</th>'] + [f'<th>{h}</th>' for h in headers]
                    header_row_html = '\n                                            '.join(header_cells)

                    excel_detail_row = f'''                                    <tr id="{tag_excel_detail_id}" class="page-tag-excel-row hidden">
                                        <td colspan="5">
                                            <div class="page-tag-excel-panel">
                                                <div class="page-tag-excel-header">
                                                    <span class="page-tag-excel-title">Excel Data: {normalized_tag}</span>
                                                </div>
                                                <table class="page-tag-excel-table">
                                                    <thead>
                                                        <tr>
                                                            {header_row_html}
                                                        </tr>
                                                    </thead>
                                                    <tbody>
{detail_row_html}
                                                    </tbody>
                                                </table>
                                            </div>
                                        </td>
                                    </tr>'''
                    detail_entries.append(excel_detail_row)

            detail_rows_html = '\n'.join(detail_entries)
            detail_row = f'''                    <tr id="{detail_id}" class="page-detail-row hidden">
                        <td colspan="{page_stats_column_count}">
                            <div class="page-detail-panel">
                                <div class="page-detail-header">
                                    <span class="page-detail-title">Tags detected on this page</span>
                                    <span class="badge">{detail_count} {detail_label}</span>
                                </div>
                                <table class="tag-detail-table">
                                    <thead>
                                        <tr>
                                            <th>Detected Text</th>
                                            <th>Matched Tag</th>
                                            <th>Excel</th>
                                            <th>Highlighted</th>
                                            <th>Coloring Reason</th>
                                        </tr>
                                    </thead>
                                    <tbody>
{detail_rows_html}
                                    </tbody>
                                </table>
                            </div>
                        </td>
                    </tr>'''
        else:
            detail_row = ''

        page_stats_rows.append(row_html + ('\n' + detail_row if detail_row else ''))

    page_stats_rows_html = '\n'.join(page_stats_rows)

    # Build optional sections with collapsible support
    duplicates_section = ''
    if duplicate_count > 0:
        duplicates_section = f'''
        <div class="section collapsible" id="duplicatesSection">
            <h2 onclick="toggleSection('duplicatesSection')">
                <span>
                    <span style="color: #dc2626;">⚠</span> Duplicate Tags <span class="badge">{duplicate_count}</span>
                </span>
                <span class="toggle-icon">▼</span>
            </h2>
            <div class="section-content">
                <table id="duplicatesTable">
                    <thead>
                        <tr>
                            <th onclick="sortTable('duplicatesTable', 0)">Tag <span class="sort-icon">↕</span></th>
                            <th onclick="sortTable('duplicatesTable', 1)">Excel Rows <span class="sort-icon">↕</span></th>
                            <th onclick="sortTable('duplicatesTable', 2)">Count <span class="sort-icon">↕</span></th>
                        </tr>
                    </thead>
                    <tbody>
{duplicates_rows_html}
                    </tbody>
                </table>
            </div>
        </div>
        '''

    warnings_section = ''
    if warning_count > 0:
        warnings_section = f'''
        <div class="section">
            <h2>Validation Warnings <span class="badge">{warning_count}</span></h2>
            <table id="warningsTable">
                <thead>
                    <tr>
                        <th onclick="sortTable('warningsTable', 0)">Tag <span class="sort-icon">↕</span></th>
                        <th onclick="sortTable('warningsTable', 1)">Excel Row <span class="sort-icon">↕</span></th>
                        <th onclick="sortTable('warningsTable', 2)">Warning <span class="sort-icon">↕</span></th>
                    </tr>
                </thead>
                <tbody>
{warnings_rows_html}
                </tbody>
            </table>
        </div>
        '''

    # Build unmatched PDF tags section (tags in PDF but not in Excel)
    unmatched_section = ''
    if unmatched_count > 0:
        unmatched_section = f'''
        <div class="section collapsible collapsed" id="unmatchedSection">
            <h2 onclick="toggleSection('unmatchedSection')">
                <span>
                    <span style="color: #8b5cf6;">&#128203;</span> Tags Found in PDF (Not in Excel) <span class="badge" style="background: #8b5cf6;" id="unmatchedBadge">{unmatched_count}</span>
                </span>
                <span class="toggle-icon">▼</span>
            </h2>
            <div class="section-content">
                <p style="color: #6b7280; margin-bottom: 12px; font-size: 14px;">
                    These tags were detected in the PDF document but do not appear in the Excel component list.
                </p>
                <div class="unmatched-filter-controls">
                    <div class="filter-group">
                        <label>Exclude by Tag Part</label>
                        <select id="unmatchedPartSelect">
                            <option value="1">Part 1 (e.g. 101)</option>
                            <option value="2" selected>Part 2 (e.g. RB)</option>
                            <option value="3">Part 3 (e.g. 07)</option>
                            <option value="4">Part 4</option>
                            <option value="5">Part 5</option>
                        </select>
                    </div>
                    <div class="filter-group">
                        <label>Values to Exclude (comma-separated)</label>
                        <input type="text" id="unmatchedExcludeInput" placeholder="e.g. RB,PB,XV">
                    </div>
                    <div class="filter-group">
                        <button class="btn purple" onclick="applyUnmatchedFilter()">Apply Filter</button>
                    </div>
                    <div class="filter-group">
                        <button class="btn secondary" onclick="clearUnmatchedFilter()">Clear Filter</button>
                    </div>
                    <span class="excluded-count" id="unmatchedFilterStatus"></span>
                </div>
                <table id="unmatchedTable">
                    <thead>
                        <tr>
                            <th onclick="sortTable('unmatchedTable', 0)">Tag <span class="sort-icon">↕</span></th>
                            <th onclick="sortTable('unmatchedTable', 1)">Pages <span class="sort-icon">↕</span></th>
                            <th onclick="sortTable('unmatchedTable', 2)">Occurrences <span class="sort-icon">↕</span></th>
                        </tr>
                    </thead>
                    <tbody>
{unmatched_rows_html}
                    </tbody>
                </table>
                <div class="unmatched-export-bar">
                    <button class="btn purple" onclick="exportUnmatchedToExcel()">Export Unmatched Tags (Excel)</button>
                    <span class="excluded-count" id="unmatchedExportInfo"></span>
                </div>
            </div>
        </div>
        '''

    # Build page statistics section for dedicated page view
    page_stats_section_html = ''
    if page_stats_count > 0:
        page_summary_html = ''
        if most_tagged_page:
            most_tagged_label = _format_page_label(most_tagged_page, include_source_pdf)
            page_summary_html = f'''
                <div class="grid grid-cols-2 md:grid-cols-4 gap-4 mb-4">
                    <div class="stat-mini-card">
                        <div class="stat-mini-label">Pages with Tags</div>
                        <div class="stat-mini-value">{total_pages_with_tags}</div>
                    </div>
                    <div class="stat-mini-card">
                        <div class="stat-mini-label">Pages with Colored Tags</div>
                        <div class="stat-mini-value">{total_pages_with_colored}</div>
                    </div>
                    <div class="stat-mini-card">
                        <div class="stat-mini-label">Most Tagged Page</div>
                        <div class="stat-mini-value">{most_tagged_label} ({most_tagged_page['total_count']} tags)</div>
                    </div>
                    <div class="stat-mini-card">
                        <div class="stat-mini-label">Avg Tags per Page</div>
                        <div class="stat-mini-value">{avg_tags_per_page:.1f}</div>
                    </div>
                </div>
            '''

        if include_source_pdf:
            page_stats_header_html = '''
                        <tr>
                            <th onclick="sortTable('pageStatsTable', 0)">Page # <span class="sort-icon">↕</span></th>
                            <th onclick="sortTable('pageStatsTable', 1)">PDF File <span class="sort-icon">↕</span></th>
                            <th onclick="sortTable('pageStatsTable', 2)">Bookmark <span class="sort-icon">↕</span></th>
                            <th onclick="sortTable('pageStatsTable', 3)">Colored Tags <span class="sort-icon">↕</span></th>
                            <th onclick="sortTable('pageStatsTable', 4)">Total Tags <span class="sort-icon">↕</span></th>
                            <th onclick="sortTable('pageStatsTable', 5)">Percentage <span class="sort-icon">↕</span></th>
                        </tr>'''
        else:
            page_stats_header_html = '''
                        <tr>
                            <th onclick="sortTable('pageStatsTable', 0)">Page # <span class="sort-icon">↕</span></th>
                            <th onclick="sortTable('pageStatsTable', 1)">Bookmark <span class="sort-icon">↕</span></th>
                            <th onclick="sortTable('pageStatsTable', 2)">Colored Tags <span class="sort-icon">↕</span></th>
                            <th onclick="sortTable('pageStatsTable', 3)">Total Tags <span class="sort-icon">↕</span></th>
                            <th onclick="sortTable('pageStatsTable', 4)">Percentage <span class="sort-icon">↕</span></th>
                        </tr>'''

        page_stats_section_html = f'''
        <div class="page-panel-container print-page-break">
            <div class="panel page-panel" id="pageStatsPanel">
                <div class="panel-header">
                    <h2><span style="color: #0ea5e9;">📊</span> Page Statistics</h2>
                    <span class="badge">{page_stats_count} pages</span>
                </div>
                {page_summary_html}
                <div class="nested-collapsible collapsed" id="pageStatsSection">
                    <button type="button" class="toggle-link" onclick="toggleSection('pageStatsSection')">
                        Detailed Breakdown <span class="toggle-icon">▼</span>
                    </button>
                    <div class="nested-content">
                        <table id="pageStatsTable">
                            <thead>
{page_stats_header_html}
                            </thead>
                            <tbody>
{page_stats_rows_html}
                            </tbody>
                        </table>
                    </div>
                </div>
            </div>
        </div>
        '''

    # Build settings footer
    settings_html = ''
    if settings:
        settings_list = []
        if "tag_column" in settings:
            settings_list.append(f'<li>Tag Column: {settings.get("tag_column", "N/A")}</li>')
        if "header_row" in settings:
            settings_list.append(f'<li>Header Row: {settings.get("header_row", "N/A")}</li>')
        if "watermark_enabled" in settings:
            status = 'Enabled' if settings.get("watermark_enabled") else 'Disabled'
            settings_list.append(f'<li>Watermark: {status}</li>')
        if "annotate_excel" in settings:
            status = 'Enabled' if settings.get("annotate_excel") else 'Disabled'
            settings_list.append(f'<li>Excel Annotation: {status}</li>')
        settings_html = '\n                '.join(settings_list)

    # JSON data for JavaScript
    report_data_json = json.dumps({
        'found': found,
        'not_found': not_found,
        'duplicates': duplicates_list,
        'validation_warnings': validation_warnings,
        'page_stats': page_stats_list,
        'unmatched_pdf_tags': unmatched_pdf_tags,
        'pdf_filenames': pdf_filenames if isinstance(pdf_filenames, list) else [str(pdf_filenames)]
    })

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>PID Annotator Processing Report</title>
    <style>
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}

        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Arial, sans-serif;
            line-height: 1.6;
            color: #333;
            background: #f5f5f5;
            padding: 20px;
        }}

        .container {{
            margin: 0 auto;
            background: white;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
            overflow: hidden;
        }}

        header {{
            background: linear-gradient(135deg, #2563eb 0%, #1d4ed8 100%);
            color: white;
            padding: 30px;
        }}

        header h1 {{
            font-size: 28px;
            margin-bottom: 8px;
        }}

        header .metadata {{
            opacity: 0.9;
            font-size: 14px;
        }}

        .overview-panels {{
            display: flex;
            flex-direction: column;
            gap: 20px;
            padding: 30px;
            border-bottom: 1px solid #e5e7eb;
        }}

        .page-panel-container {{
            padding: 30px;
            border-bottom: 1px solid #e5e7eb;
        }}

        .page-panel-container .panel {{
            margin: 0;
        }}

        .print-page-break {{
            margin-top: 30px;
        }}

        .panel {{
            background: #f9fafb;
            border: 1px solid #e5e7eb;
            border-radius: 12px;
            padding: 24px;
            display: flex;
            flex-direction: column;
            gap: 16px;
        }}

        .panel-header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 12px;
        }}

        .panel-header h2 {{
            font-size: 20px;
            margin: 0;
            color: #111827;
            display: flex;
            align-items: center;
            gap: 8px;
        }}

        .summary {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
        }}

        .stat-card {{
            background: white;
            padding: 20px;
            border-radius: 8px;
            border-left: 4px solid #2563eb;
        }}

        .stat-card.success {{ border-left-color: #16a34a; }}
        .stat-card.warning {{ border-left-color: #f59e0b; }}
        .stat-card.error {{ border-left-color: #dc2626; }}
        .stat-card.info {{ border-left-color: #0ea5e9; }}

        .stat-card .label {{
            font-size: 12px;
            color: #6b7280;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            margin-bottom: 4px;
        }}

        .stat-card .value {{
            font-size: 32px;
            font-weight: bold;
            color: #111827;
        }}

        .stat-card .subtext {{
            font-size: 13px;
            color: #6b7280;
            margin-top: 4px;
        }}

        .controls {{
            padding: 20px 30px;
            background: #f9fafb;
            border-bottom: 1px solid #e5e7eb;
            display: flex;
            gap: 15px;
            flex-wrap: wrap;
            align-items: center;
        }}

        .search-box {{
            flex: 1;
            min-width: 250px;
            padding: 10px 15px;
            border: 1px solid #d1d5db;
            border-radius: 6px;
            font-size: 14px;
        }}

        .search-box:focus {{
            outline: none;
            border-color: #2563eb;
            box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.1);
        }}

        .btn {{
            padding: 10px 20px;
            background: #2563eb;
            color: white;
            border: none;
            border-radius: 6px;
            cursor: pointer;
            font-size: 14px;
            font-weight: 500;
            transition: background 0.2s;
        }}

        .btn:hover {{ background: #1d4ed8; }}
        .btn.secondary {{ background: #6b7280; }}
        .btn.secondary:hover {{ background: #4b5563; }}
        .btn.purple {{ background: #8b5cf6; }}
        .btn.purple:hover {{ background: #7c3aed; }}

        .unmatched-filter-controls {{
            background: #f5f3ff;
            border: 1px solid #e0d5f5;
            border-radius: 8px;
            padding: 16px 20px;
            margin-bottom: 16px;
            display: flex;
            flex-wrap: wrap;
            gap: 12px;
            align-items: flex-end;
        }}

        .filter-group {{
            display: flex;
            flex-direction: column;
            gap: 4px;
        }}

        .filter-group label {{
            font-size: 12px;
            font-weight: 600;
            color: #6b7280;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}

        .filter-group select,
        .filter-group input {{
            padding: 8px 12px;
            border: 1px solid #d1d5db;
            border-radius: 6px;
            font-size: 14px;
            background: white;
        }}

        .filter-group select:focus,
        .filter-group input:focus {{
            outline: none;
            border-color: #8b5cf6;
            box-shadow: 0 0 0 3px rgba(139, 92, 246, 0.1);
        }}

        .filter-group input {{
            min-width: 220px;
        }}

        .excluded-count {{
            font-size: 13px;
            color: #8b5cf6;
            font-weight: 500;
            padding: 8px 0;
        }}

        .unmatched-export-bar {{
            display: flex;
            gap: 10px;
            align-items: center;
            margin-top: 12px;
            padding-top: 12px;
            border-top: 1px solid #e0d5f5;
        }}

        .section {{
            padding: 30px;
        }}

        .section h2 {{
            font-size: 20px;
            margin-bottom: 15px;
            color: #111827;
            display: flex;
            align-items: center;
            gap: 10px;
        }}

        .badge {{
            display: inline-block;
            padding: 2px 8px;
            border-radius: 12px;
            font-size: 12px;
            font-weight: 600;
            background: #e5e7eb;
            color: #374151;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 15px;
        }}

        th {{
            background: #f3f4f6;
            padding: 12px 15px;
            text-align: left;
            font-weight: 600;
            font-size: 13px;
            color: #374151;
            border-bottom: 2px solid #e5e7eb;
            cursor: pointer;
            user-select: none;
        }}

        th:hover {{ background: #e5e7eb; }}

        th .sort-icon {{
            margin-left: 5px;
            opacity: 0.3;
            font-size: 10px;
        }}

        th.sorted .sort-icon {{ opacity: 1; }}

        td {{
            padding: 12px 15px;
            border-bottom: 1px solid #e5e7eb;
            font-size: 14px;
        }}

        tr:hover {{ background: #f9fafb; }}

        .tag-name {{
            font-family: 'Courier New', monospace;
            font-weight: 600;
            color: #2563eb;
        }}

        .page-list {{
            font-size: 13px;
            color: #6b7280;
        }}

        .bookmark-list {{
            font-size: 13px;
            color: #059669;
            font-style: italic;
        }}

        .excel-row {{
            font-size: 13px;
            color: #6b7280;
        }}

        .warning-text {{
            color: #dc2626;
            font-size: 13px;
        }}

        .page-number {{
            font-weight: 600;
            color: #2563eb;
        }}

        .bookmark-title {{
            font-size: 13px;
            color: #059669;
            font-style: italic;
            max-width: 300px;
            white-space: nowrap;
            overflow: hidden;
            text-overflow: ellipsis;
        }}

        .stat-mini-card {{
            background: #f9fafb;
            padding: 15px;
            border-radius: 8px;
            border-left: 3px solid #0ea5e9;
        }}

        .stat-mini-label {{
            font-size: 11px;
            color: #6b7280;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            margin-bottom: 4px;
        }}

        .stat-mini-value {{
            font-size: 18px;
            font-weight: bold;
            color: #111827;
        }}

        .percentage-bar-container {{
            width: 100px;
            height: 12px;
            background: #e5e7eb;
            border-radius: 6px;
            overflow: hidden;
        }}

        .percentage-bar {{
            height: 100%;
            background: linear-gradient(90deg, #10b981 0%, #059669 100%);
            transition: width 0.3s ease;
        }}

        .percentage-row {{
            display: flex;
            align-items: center;
            gap: 0.5rem;
        }}

        .percentage-value {{
            font-size: 13px;
            color: #374151;
            font-weight: 500;
        }}

        .page-summary-cell {{
            display: flex;
            align-items: center;
            gap: 8px;
        }}

        .page-detail-toggle {{
            width: 28px;
            height: 28px;
            border: none;
            border-radius: 9999px;
            background: #e0e7ff;
            color: #312e81;
            cursor: pointer;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            transition: background 0.2s ease, transform 0.2s ease;
        }}

        .page-detail-toggle:hover {{
            background: #c7d2fe;
        }}

        .page-detail-toggle:focus {{
            outline: none;
            box-shadow: 0 0 0 3px rgba(99, 102, 241, 0.25);
        }}

        .page-detail-toggle .chevron {{
            font-size: 12px;
            transition: transform 0.2s ease;
        }}

        .page-detail-toggle.open .chevron {{
            transform: rotate(180deg);
        }}

        .page-detail-row td {{
            background: #f8fafc;
            padding: 0;
        }}

        .page-detail-panel {{
            background: #ffffff;
            padding: 16px 20px;
            border-top: 1px solid #e2e8f0;
        }}

        .page-detail-header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 12px;
            margin-bottom: 12px;
        }}

        .page-detail-title {{
            font-size: 15px;
            font-weight: 600;
            color: #0f172a;
        }}

        .tag-detail-table {{
            width: 100%;
            border-collapse: collapse;
        }}

        .tag-detail-table th,
        .tag-detail-table td {{
            padding: 10px 12px;
            border-bottom: 1px solid #e5e7eb;
            font-size: 13px;
            text-align: left;
        }}

        .tag-detail-table th {{
            background: #eff6ff;
            color: #1e3a8a;
            font-weight: 600;
        }}

        .tag-detail-table tr:last-child td {{
            border-bottom: none;
        }}

        .detail-tag {{
            font-family: 'Courier New', monospace;
            font-weight: 600;
            color: #1d4ed8;
        }}

        /* Nested page tag Excel detail styles */
        .page-tag-cell {{
            display: flex;
            align-items: center;
            gap: 6px;
        }}

        .page-tag-excel-toggle {{
            width: 20px;
            height: 20px;
            border: none;
            border-radius: 4px;
            background: #dbeafe;
            color: #1e40af;
            cursor: pointer;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            transition: background 0.2s ease;
            font-size: 10px;
        }}

        .page-tag-excel-toggle:hover {{
            background: #bfdbfe;
        }}

        .page-tag-excel-toggle:focus {{
            outline: none;
            box-shadow: 0 0 0 2px rgba(59, 130, 246, 0.25);
        }}

        .page-tag-excel-toggle .chevron-small {{
            font-size: 10px;
            transition: transform 0.2s ease;
        }}

        .page-tag-excel-toggle.open .chevron-small {{
            transform: rotate(180deg);
        }}

        .page-tag-excel-row td {{
            background: #f0f9ff;
            padding: 0;
        }}

        .page-tag-excel-panel {{
            background: #ffffff;
            padding: 12px 16px;
            border-top: 1px solid #bfdbfe;
            border-left: 3px solid #3b82f6;
            margin: 4px 0;
        }}

        .page-tag-excel-header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 8px;
            margin-bottom: 8px;
        }}

        .page-tag-excel-title {{
            font-size: 13px;
            font-weight: 600;
            color: #1e40af;
        }}

        .page-tag-excel-table {{
            width: 100%;
            border-collapse: collapse;
        }}

        .page-tag-excel-table th,
        .page-tag-excel-table td {{
            padding: 8px 10px;
            border-bottom: 1px solid #e5e7eb;
            font-size: 12px;
            text-align: left;
        }}

        .page-tag-excel-table th {{
            background: #eff6ff;
            color: #1e3a8a;
            font-weight: 600;
        }}

        .page-tag-excel-table tr:last-child td {{
            border-bottom: none;
        }}

        .status-badge {{
            display: inline-flex;
            align-items: center;
            justify-content: center;
            padding: 4px 10px;
            border-radius: 9999px;
            font-size: 12px;
            font-weight: 600;
            letter-spacing: 0.3px;
        }}

        .status-success {{
            background: #dcfce7;
            color: #15803d;
        }}

        .status-warning {{
            background: #fee2e2;
            color: #b91c1c;
        }}

        .status-muted {{
            background: #e5e7eb;
            color: #374151;
        }}

        /* Coloring reason styles */
        .coloring-reason {{
            display: inline-flex;
            align-items: center;
            justify-content: center;
            padding: 4px 10px;
            border-radius: 6px;
            font-size: 12px;
            font-weight: 500;
            letter-spacing: 0.3px;
        }}

        .reason-user {{
            background: #dbeafe;
            color: #1e40af;
            border-left: 3px solid #2563eb;
        }}

        .reason-auto {{
            background: #d1fae5;
            color: #065f46;
            border-left: 3px solid #10b981;
        }}

        .reason-custom {{
            background: #fef3c7;
            color: #92400e;
            border-left: 3px solid #f59e0b;
        }}

        .reason-none {{
            background: #f3f4f6;
            color: #6b7280;
            border-left: 3px solid #9ca3af;
        }}

        .reason-default {{
            background: #ede9fe;
            color: #5b21b6;
            border-left: 3px solid #8b5cf6;
        }}

        /* Duplicate detail panel styles */
        .duplicate-summary-cell {{
            display: flex;
            align-items: center;
            gap: 8px;
        }}

        .duplicate-detail-toggle {{
            width: 28px;
            height: 28px;
            border: none;
            border-radius: 9999px;
            background: #fee2e2;
            color: #7f1d1d;
            cursor: pointer;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            transition: background 0.2s ease, transform 0.2s ease;
        }}

        .duplicate-detail-toggle:hover {{
            background: #fecaca;
        }}

        .duplicate-detail-toggle:focus {{
            outline: none;
            box-shadow: 0 0 0 3px rgba(220, 38, 38, 0.25);
        }}

        .duplicate-detail-toggle .chevron {{
            font-size: 12px;
            transition: transform 0.2s ease;
        }}

        .duplicate-detail-toggle.open .chevron {{
            transform: rotate(180deg);
        }}

        .duplicate-detail-row td {{
            background: #fef2f2;
            padding: 0;
        }}

        .duplicate-detail-panel {{
            background: #ffffff;
            padding: 16px 20px;
            border-top: 1px solid #fecaca;
        }}

        .duplicate-detail-header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 12px;
            margin-bottom: 12px;
        }}

        .duplicate-detail-title {{
            font-size: 15px;
            font-weight: 600;
            color: #7f1d1d;
        }}

        .duplicate-data-table {{
            width: 100%;
            border-collapse: collapse;
        }}

        .duplicate-data-table th,
        .duplicate-data-table td {{
            padding: 10px 12px;
            border-bottom: 1px solid #e5e7eb;
            font-size: 13px;
            text-align: left;
        }}

        .duplicate-data-table th {{
            background: #fef2f2;
            color: #991b1b;
            font-weight: 600;
        }}

        .duplicate-data-table tr:last-child td {{
            border-bottom: none;
        }}

        /* Not Found detail panel styles */
        .not-found-data-table {{
            width: 100%;
            border-collapse: collapse;
        }}

        .not-found-data-table th,
        .not-found-data-table td {{
            padding: 10px 12px;
            border-bottom: 1px solid #e5e7eb;
            font-size: 13px;
            text-align: left;
        }}

        .not-found-data-table th {{
            background: #fffbeb;
            color: #92400e;
            font-weight: 600;
        }}

        .not-found-data-table tr:last-child td {{
            border-bottom: none;
        }}

        /* Continue Not Found detail panel styles */
        .not-found-summary-cell {{
            display: flex;
            align-items: center;
            gap: 8px;
        }}

        .not-found-detail-toggle {{
            width: 28px;
            height: 28px;
            border: none;
            border-radius: 9999px;
            background: #fef3c7;
            color: #92400e;
            cursor: pointer;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            transition: background 0.2s ease, transform 0.2s ease;
        }}

        .not-found-detail-toggle:hover {{
            background: #fde68a;
        }}

        .not-found-detail-toggle:focus {{
            outline: none;
            box-shadow: 0 0 0 3px rgba(180, 83, 9, 0.25);
        }}

        .not-found-detail-toggle .chevron {{
            font-size: 12px;
            transition: transform 0.2s ease;
        }}

        .not-found-detail-toggle.open .chevron {{
            transform: rotate(180deg);
        }}

        .not-found-detail-row td {{
            background: #fffbeb;
            padding: 0;
        }}

        .not-found-detail-panel {{
            background: #ffffff;
            padding: 16px 20px;
            border-top: 1px solid #fde68a;
        }}

        .not-found-detail-header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 12px;
            margin-bottom: 12px;
        }}

        .not-found-detail-title {{
            font-size: 15px;
            font-weight: 600;
            color: #92400e;
        }}

        .not-found-detail-content {{
            display: flex;
            flex-direction: column;
            gap: 10px;
        }}

        .detail-item {{
            padding: 10px;
            background: #faf5ff;
            border-radius: 6px;
            font-size: 13px;
        }}

        .detail-item strong {{
            color: #92400e;
            margin-right: 8px;
        }}

        .tag-value {{
            font-family: 'Courier New', monospace;
            font-weight: 600;
            color: #b45309;
        }}

        .reason-value {{
            color: #ea580c;
            font-weight: 500;
        }}

        /* Found detail panel styles */
        .found-summary-cell {{
            display: flex;
            align-items: center;
            gap: 8px;
        }}

        .found-detail-toggle {{
            width: 28px;
            height: 28px;
            border: none;
            border-radius: 9999px;
            background: #d1fae5;
            color: #065f46;
            cursor: pointer;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            transition: background 0.2s ease, transform 0.2s ease;
        }}

        .found-detail-toggle:hover {{
            background: #a7f3d0;
        }}

        .found-detail-toggle:focus {{
            outline: none;
            box-shadow: 0 0 0 3px rgba(16, 185, 129, 0.25);
        }}

        .found-detail-toggle .chevron {{
            font-size: 12px;
            transition: transform 0.2s ease;
        }}

        .found-detail-toggle.open .chevron {{
            transform: rotate(180deg);
        }}

        .found-detail-row td {{
            background: #ecfdf5;
            padding: 0;
        }}

        .found-detail-panel {{
            background: #ffffff;
            padding: 16px 20px;
            border-top: 1px solid #a7f3d0;
        }}

        .found-detail-header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 12px;
            margin-bottom: 12px;
        }}

        .found-detail-title {{
            font-size: 15px;
            font-weight: 600;
            color: #065f46;
        }}

        .found-data-table {{
            width: 100%;
            border-collapse: collapse;
        }}

        .found-data-table th,
        .found-data-table td {{
            padding: 10px 12px;
            border-bottom: 1px solid #e5e7eb;
            font-size: 13px;
            text-align: left;
        }}

        .found-data-table th {{
            background: #d1fae5;
            color: #065f46;
            font-weight: 600;
        }}

        .found-data-table tr:last-child td {{
            border-bottom: none;
        }}

        /* Warning detail panel styles */
        .warning-summary-cell {{
            display: flex;
            align-items: center;
            gap: 8px;
        }}

        .warning-detail-toggle {{
            width: 28px;
            height: 28px;
            border: none;
            border-radius: 9999px;
            background: #fed7aa;
            color: #7c2d12;
            cursor: pointer;
            display: inline-flex;
            align-items: center;
            justify-content: center;
            transition: background 0.2s ease, transform 0.2s ease;
        }}

        .warning-detail-toggle:hover {{
            background: #fdba74;
        }}

        .warning-detail-toggle:focus {{
            outline: none;
            box-shadow: 0 0 0 3px rgba(249, 115, 22, 0.25);
        }}

        .warning-detail-toggle .chevron {{
            font-size: 12px;
            transition: transform 0.2s ease;
        }}

        .warning-detail-toggle.open .chevron {{
            transform: rotate(180deg);
        }}

        .warning-detail-row td {{
            background: #ffedd5;
            padding: 0;
        }}

        .warning-detail-panel {{
            background: #ffffff;
            padding: 16px 20px;
            border-top: 1px solid #fdba74;
        }}

        .warning-detail-header {{
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 12px;
            margin-bottom: 12px;
        }}

        .warning-detail-title {{
            font-size: 15px;
            font-weight: 600;
            color: #7c2d12;
        }}

        .warning-data-table {{
            width: 100%;
            border-collapse: collapse;
        }}

        .warning-data-table th,
        .warning-data-table td {{
            padding: 10px 12px;
            border-bottom: 1px solid #e5e7eb;
            font-size: 13px;
            text-align: left;
        }}

        .warning-data-table th {{
            background: #ffedd5;
            color: #9a3412;
            font-weight: 600;
        }}

        .warning-data-table tr:last-child td {{
            border-bottom: none;
        }}

        .grid {{
            display: grid;
        }}

        .grid-cols-2 {{
            grid-template-columns: repeat(2, 1fr);
        }}

        .gap-4 {{
            gap: 1rem;
        }}

        .mb-4 {{
            margin-bottom: 1rem;
        }}

        @media (min-width: 768px) {{
            .md\\:grid-cols-4 {{
                grid-template-columns: repeat(4, 1fr);
            }}
        }}

        footer {{
            padding: 30px;
            background: #f9fafb;
            border-top: 1px solid #e5e7eb;
        }}

        footer h3 {{
            font-size: 16px;
            margin-bottom: 10px;
            color: #111827;
        }}

        footer ul {{
            list-style: none;
            font-size: 14px;
            color: #6b7280;
        }}

        footer li {{
            padding: 4px 0;
        }}

        .hidden {{ display: none; }}

        /* Collapsible section styles */
        .section.collapsible {{
            border: 1px solid #e5e7eb;
            border-radius: 8px;
            margin-bottom: 20px;
            overflow: hidden;
        }}

        .section.collapsible h2 {{
            cursor: pointer;
            padding: 20px 30px;
            margin: 0;
            background: #f9fafb;
            border-bottom: 1px solid #e5e7eb;
            user-select: none;
            display: flex;
            align-items: center;
            justify-content: space-between;
        }}

        .section.collapsible h2:hover {{
            background: #f3f4f6;
        }}

        .section.collapsible h2 .toggle-icon {{
            font-size: 20px;
            transition: transform 0.3s;
        }}

        .section.collapsible.collapsed h2 .toggle-icon {{
            transform: rotate(-90deg);
        }}

        .section.collapsible .section-content {{
            padding: 0 30px 30px 30px;
            transition: max-height 0.3s ease-out;
        }}

        .section.collapsible.collapsed .section-content {{
            display: none;
        }}

        .nested-collapsible {{
            border-top: 1px solid #e5e7eb;
            padding-top: 16px;
        }}

        .nested-collapsible .toggle-link {{
            background: transparent;
            border: none;
            color: #2563eb;
            font-weight: 600;
            cursor: pointer;
            display: flex;
            align-items: center;
            gap: 8px;
            font-size: 14px;
            padding: 0;
        }}

        .nested-collapsible .toggle-icon {{
            font-size: 16px;
            transition: transform 0.3s;
        }}

        .nested-collapsible.collapsed .toggle-icon {{
            transform: rotate(-90deg);
        }}

        .nested-content {{
            margin-top: 16px;
        }}

        .nested-collapsible.collapsed .nested-content {{
            display: none;
        }}

        @media print {{
            body {{ padding: 0; background: white; }}
            .controls {{ display: none; }}
            .container {{ box-shadow: none; }}
            tr:hover {{ background: transparent; }}
            .section.collapsible.collapsed .section-content {{
                display: block !important;
            }}
            .toggle-icon {{ display: none; }}
            .nested-collapsible .toggle-link {{ display: none; }}
            .nested-collapsible .nested-content {{
                display: block !important;
            }}
            .print-page-break {{
                page-break-before: always;
                margin-top: 0;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>PID Annotation Processing Report</h1>
            <div class="metadata">
                Processed: {process_date} | PDF: {pdf_list} | Excel: {excel_filename}
            </div>
        </header>

        <div class="overview-panels">
            <div class="panel tag-panel">
                <div class="panel-header">
                    <h2><span style="color: #2563eb;">🏷️</span> Tag Statistics</h2>
                </div>
                <div class="summary">
                    <div class="stat-card">
                        <div class="label">Total Tags</div>
                        <div class="value">{total_tags}</div>
                        <div class="subtext">in Excel file</div>
                    </div>
                    <div class="stat-card success">
                        <div class="label">Found</div>
                        <div class="value">{found_count}</div>
                        <div class="subtext">{found_percent:.1f}% success rate</div>
                    </div>
                    <div class="stat-card warning">
                        <div class="label">Not Found</div>
                        <div class="value">{not_found_count}</div>
                        <div class="subtext">{not_found_percent:.1f}% missing</div>
                    </div>
                    <div class="stat-card {'info' if duplicate_count > 0 else ''}">
                        <div class="label">Duplicates</div>
                        <div class="value">{duplicate_count}</div>
                        <div class="subtext">{'tags appear multiple times' if duplicate_count > 0 else 'no duplicates detected'}</div>
                    </div>
                </div>
            </div>
        </div>
        {page_stats_section_html}

        <div class="controls">
            <input type="text" id="searchBox" class="search-box" placeholder="Search tags...">
            <button class="btn" onclick="exportFilteredToExcel()">Export Filtered Data (Excel)</button>
            <button class="btn secondary" onclick="window.print()">Print Report</button>
        </div>

{duplicates_section}

        <div class="section collapsible collapsed" id="notFoundSection">
            <h2 onclick="toggleSection('notFoundSection')">
                <span>
                    <span style="color: #f59e0b;">⊘</span> Not Found Tags <span class="badge">{not_found_count}</span>
                </span>
                <span class="toggle-icon">▼</span>
            </h2>
            <div class="section-content">
                <table id="notFoundTable">
                    <thead>
                        <tr>
                            <th onclick="sortTable('notFoundTable', 0)">Tag <span class="sort-icon">↕</span></th>
                            <th onclick="sortTable('notFoundTable', 1)">Excel Row <span class="sort-icon">↕</span></th>
                            <th onclick="sortTable('notFoundTable', 2)">Reason <span class="sort-icon">↕</span></th>
                        </tr>
                    </thead>
                    <tbody>
{not_found_rows_html}
                    </tbody>
                </table>
            </div>
        </div>

        <div class="section collapsible collapsed" id="foundSection">
            <h2 onclick="toggleSection('foundSection')">
                <span>
                    <span style="color: #16a34a;">✓</span> Found Tags <span class="badge">{found_count}</span>
                </span>
                <span class="toggle-icon">▼</span>
            </h2>
            <div class="section-content">
                <table id="foundTable">
                    <thead>
                        <tr>
                            <th onclick="sortTable('foundTable', 0)">Tag <span class="sort-icon">↕</span></th>
                            <th onclick="sortTable('foundTable', 1)">Pages <span class="sort-icon">↕</span></th>
                            <th onclick="sortTable('foundTable', 2)">Bookmarks <span class="sort-icon">↕</span></th>
                            <th onclick="sortTable('foundTable', 3)">Occurrences <span class="sort-icon">↕</span></th>
                            <th onclick="sortTable('foundTable', 4)">Excel Row <span class="sort-icon">↕</span></th>
                        </tr>
                    </thead>
                    <tbody>
{found_rows_html}
                    </tbody>
                </table>
            </div>
        </div>

{unmatched_section}

{warnings_section}

        <footer>
            <h3>Processing Settings</h3>
            <ul>
                {settings_html}
            </ul>
        </footer>
    </div>

    <script src="https://cdn.sheetjs.com/xlsx-0.20.1/package/dist/xlsx.full.min.js"></script>
    <script>
        // Report data embedded as JSON
        const reportData = {report_data_json};

        // Toggle collapsible sections
        function toggleSection(sectionId) {{
            const section = document.getElementById(sectionId);
            if (section) {{
                section.classList.toggle('collapsed');
            }}
        }}

        function togglePageDetail(detailId, trigger) {{
            const row = document.getElementById(detailId);
            if (!row) {{
                return false;
            }}

            const isHidden = row.classList.toggle('hidden');

            if (isHidden) {{
                row.style.display = 'none';
            }} else {{
                row.style.display = '';
            }}

            if (trigger) {{
                trigger.classList.toggle('open', !isHidden);
                trigger.setAttribute('aria-expanded', String(!isHidden));
            }}

            return false;
        }}

        function toggleDuplicateDetail(detailId, trigger) {{
            const row = document.getElementById(detailId);
            if (!row) {{
                return false;
            }}

            const isHidden = row.classList.toggle('hidden');

            if (isHidden) {{
                row.style.display = 'none';
            }} else {{
                row.style.display = '';
            }}

            if (trigger) {{
                trigger.classList.toggle('open', !isHidden);
                trigger.setAttribute('aria-expanded', String(!isHidden));
            }}

            return false;
        }}

        function toggleNotFoundDetail(detailId, trigger) {{
            const row = document.getElementById(detailId);
            if (!row) {{
                return false;
            }}

            const isHidden = row.classList.toggle('hidden');

            if (isHidden) {{
                row.style.display = 'none';
            }} else {{
                row.style.display = '';
            }}

            if (trigger) {{
                trigger.classList.toggle('open', !isHidden);
                trigger.setAttribute('aria-expanded', String(!isHidden));
            }}

            return false;
        }}

        function toggleFoundDetail(detailId, trigger) {{
            const row = document.getElementById(detailId);
            if (!row) {{
                return false;
            }}

            const isHidden = row.classList.toggle('hidden');

            if (isHidden) {{
                row.style.display = 'none';
            }} else {{
                row.style.display = '';
            }}

            if (trigger) {{
                trigger.classList.toggle('open', !isHidden);
                trigger.setAttribute('aria-expanded', String(!isHidden));
            }}

            return false;
        }}

        function toggleWarningDetail(detailId, trigger) {{
            const row = document.getElementById(detailId);
            if (!row) {{
                return false;
            }}

            const isHidden = row.classList.toggle('hidden');

            if (isHidden) {{
                row.style.display = 'none';
            }} else {{
                row.style.display = '';
            }}

            if (trigger) {{
                trigger.classList.toggle('open', !isHidden);
                trigger.setAttribute('aria-expanded', String(!isHidden));
            }}

            return false;
        }}

        function togglePageTagExcel(detailId, trigger) {{
            const row = document.getElementById(detailId);
            if (!row) {{
                return false;
            }}

            const isHidden = row.classList.toggle('hidden');

            if (isHidden) {{
                row.style.display = 'none';
            }} else {{
                row.style.display = '';
            }}

            if (trigger) {{
                trigger.classList.toggle('open', !isHidden);
                trigger.setAttribute('aria-expanded', String(!isHidden));
            }}

            return false;
        }}

        // Table sorting
        function sortTable(tableId, column) {{
            const table = document.getElementById(tableId);
            const tbody = table.querySelector('tbody');
            const allRows = Array.from(tbody.querySelectorAll('tr'));
            const th = table.querySelectorAll('th')[column];

            const isPageStatsTable = tableId === 'pageStatsTable';
            const isDuplicatesTable = tableId === 'duplicatesTable';
            const isNotFoundTable = tableId === 'notFoundTable';
            const isFoundTable = tableId === 'foundTable';
            const isWarningsTable = tableId === 'warningsTable';
            const detailRowMap = {{}};

            let rows = allRows;
            if (isPageStatsTable) {{
                rows = [];
                allRows.forEach(row => {{
                    if (row.classList.contains('page-detail-row')) {{
                        detailRowMap[row.id] = row;
                    }} else {{
                        rows.push(row);
                    }}
                }});
            }} else if (isDuplicatesTable) {{
                rows = [];
                allRows.forEach(row => {{
                    if (row.classList.contains('duplicate-detail-row')) {{
                        detailRowMap[row.id] = row;
                    }} else {{
                        rows.push(row);
                    }}
                }});
            }} else if (isNotFoundTable) {{
                rows = [];
                allRows.forEach(row => {{
                    if (row.classList.contains('not-found-detail-row')) {{
                        detailRowMap[row.id] = row;
                    }} else {{
                        rows.push(row);
                    }}
                }});
            }} else if (isFoundTable) {{
                rows = [];
                allRows.forEach(row => {{
                    if (row.classList.contains('found-detail-row')) {{
                        detailRowMap[row.id] = row;
                    }} else {{
                        rows.push(row);
                    }}
                }});
            }} else if (isWarningsTable) {{
                rows = [];
                allRows.forEach(row => {{
                    if (row.classList.contains('warning-detail-row')) {{
                        detailRowMap[row.id] = row;
                    }} else {{
                        rows.push(row);
                    }}
                }});
            }}

            // Toggle sort direction
            const isAsc = th.classList.contains('sorted-asc');

            // Remove all sorted classes
            table.querySelectorAll('th').forEach(h => {{
                h.classList.remove('sorted', 'sorted-asc', 'sorted-desc');
            }});

            // Add sorted class
            th.classList.add('sorted');
            th.classList.add(isAsc ? 'sorted-desc' : 'sorted-asc');

            // Sort rows
            rows.sort((a, b) => {{
                let aVal = a.cells[column].textContent.trim();
                let bVal = b.cells[column].textContent.trim();

                // Try to parse as numbers
                const aNum = parseFloat(aVal);
                const bNum = parseFloat(bVal);

                if (!isNaN(aNum) && !isNaN(bNum)) {{
                    return isAsc ? bNum - aNum : aNum - bNum;
                }}

                // String comparison
                return isAsc
                    ? bVal.localeCompare(aVal)
                    : aVal.localeCompare(bVal);
            }});

            // Re-append sorted rows
            tbody.innerHTML = '';

            rows.forEach(row => {{
                tbody.appendChild(row);

                if (isPageStatsTable || isDuplicatesTable || isNotFoundTable || isFoundTable || isWarningsTable) {{
                    const detailId = row.getAttribute('data-detail-id');
                    if (detailId && detailRowMap[detailId]) {{
                        tbody.appendChild(detailRowMap[detailId]);
                    }}
                }}
            }});
        }}

        // Search functionality
        document.getElementById('searchBox').addEventListener('input', function(e) {{
            const searchTerm = e.target.value.toLowerCase();
            const tables = ['foundTable', 'notFoundTable', 'duplicatesTable', 'warningsTable', 'pageStatsTable'];

            tables.forEach(tableId => {{
                const table = document.getElementById(tableId);
                if (!table) return;

                if (tableId === 'pageStatsTable') {{
                    const summaryRows = Array.from(table.querySelectorAll('tbody tr.page-summary-row'));
                    summaryRows.forEach(row => {{
                        const detailId = row.getAttribute('data-detail-id');
                        const detailRow = detailId ? document.getElementById(detailId) : null;

                        const combinedText = (row.textContent + (detailRow ? detailRow.textContent : '')).toLowerCase();
                        const isMatch = combinedText.includes(searchTerm);

                        row.style.display = isMatch ? '' : 'none';

                        if (detailRow) {{
                            detailRow.style.display = isMatch
                                ? (detailRow.classList.contains('hidden') ? 'none' : '')
                                : 'none';
                        }}
                    }});
                    return;
                }}

                if (tableId === 'duplicatesTable') {{
                    const summaryRows = Array.from(table.querySelectorAll('tbody tr.duplicate-summary-row'));
                    summaryRows.forEach(row => {{
                        const detailId = row.getAttribute('data-detail-id');
                        const detailRow = detailId ? document.getElementById(detailId) : null;

                        const combinedText = (row.textContent + (detailRow ? detailRow.textContent : '')).toLowerCase();
                        const isMatch = combinedText.includes(searchTerm);

                        row.style.display = isMatch ? '' : 'none';

                        if (detailRow) {{
                            detailRow.style.display = isMatch
                                ? (detailRow.classList.contains('hidden') ? 'none' : '')
                                : 'none';
                        }}
                    }});
                    return;
                }}

                if (tableId === 'notFoundTable') {{
                    const summaryRows = Array.from(table.querySelectorAll('tbody tr.not-found-summary-row'));
                    summaryRows.forEach(row => {{
                        const detailId = row.getAttribute('data-detail-id');
                        const detailRow = detailId ? document.getElementById(detailId) : null;

                        const combinedText = (row.textContent + (detailRow ? detailRow.textContent : '')).toLowerCase();
                        const isMatch = combinedText.includes(searchTerm);

                        row.style.display = isMatch ? '' : 'none';

                        if (detailRow) {{
                            detailRow.style.display = isMatch
                                ? (detailRow.classList.contains('hidden') ? 'none' : '')
                                : 'none';
                        }}
                    }});
                    return;
                }}

                if (tableId === 'foundTable') {{
                    const summaryRows = Array.from(table.querySelectorAll('tbody tr.found-summary-row'));
                    summaryRows.forEach(row => {{
                        const detailId = row.getAttribute('data-detail-id');
                        const detailRow = detailId ? document.getElementById(detailId) : null;

                        const combinedText = (row.textContent + (detailRow ? detailRow.textContent : '')).toLowerCase();
                        const isMatch = combinedText.includes(searchTerm);

                        row.style.display = isMatch ? '' : 'none';

                        if (detailRow) {{
                            detailRow.style.display = isMatch
                                ? (detailRow.classList.contains('hidden') ? 'none' : '')
                                : 'none';
                        }}
                    }});
                    return;
                }}

                if (tableId === 'warningsTable') {{
                    const summaryRows = Array.from(table.querySelectorAll('tbody tr.warning-summary-row'));
                    summaryRows.forEach(row => {{
                        const detailId = row.getAttribute('data-detail-id');
                        const detailRow = detailId ? document.getElementById(detailId) : null;

                        const combinedText = (row.textContent + (detailRow ? detailRow.textContent : '')).toLowerCase();
                        const isMatch = combinedText.includes(searchTerm);

                        row.style.display = isMatch ? '' : 'none';

                        if (detailRow) {{
                            detailRow.style.display = isMatch
                                ? (detailRow.classList.contains('hidden') ? 'none' : '')
                                : 'none';
                        }}
                    }});
                    return;
                }}

                const rows = table.querySelectorAll('tbody tr');
                rows.forEach(row => {{
                    const text = row.textContent.toLowerCase();
                    row.style.display = text.includes(searchTerm) ? '' : 'none';
                }});
            }});
        }});

        // Export filtered data to Excel - only exports visible/filtered rows
        function exportFilteredToExcel() {{
            if (typeof XLSX === 'undefined') {{
                alert('Excel export library not loaded. Please check your internet connection.');
                return;
            }}

            const wb = XLSX.utils.book_new();
            const timestamp = new Date().toISOString().replace(/[:.]/g, '-').slice(0, 19);
            let hasData = false;

            // Helper function to get visible rows from a table
            function getVisibleRows(tableId, headers) {{
                const table = document.getElementById(tableId);
                if (!table) return null;

                const tbody = table.querySelector('tbody');
                const rows = Array.from(tbody.querySelectorAll('tr'));
                const visibleRows = rows.filter(row => row.style.display !== 'none');

                if (visibleRows.length === 0) return null;

                const data = visibleRows.map(row => {{
                    return Array.from(row.cells).map(cell => cell.textContent.trim());
                }});

                return {{ headers, data }};
            }}

            // Collect data from all visible sections
            const sections = [
                {{ id: 'foundTable', name: 'Found Tags', headers: ['Tag', 'Pages', 'Bookmarks', 'Occurrences', 'Excel Row'] }},
                {{ id: 'notFoundTable', name: 'Not Found Tags', headers: ['Tag', 'Excel Row', 'Reason'] }},
                {{ id: 'duplicatesTable', name: 'Duplicate Tags', headers: ['Tag', 'Excel Rows', 'Count'] }},
                {{ id: 'warningsTable', name: 'Validation Warnings', headers: ['Tag', 'Excel Row', 'Warning'] }},
                {{ id: 'pageStatsTable', name: 'Page Statistics', headers: ['Page #', 'Bookmark', 'Colored Tags', 'Total Tags', 'Percentage'] }}
            ];

            sections.forEach(section => {{
                const result = getVisibleRows(section.id, section.headers);
                if (result) {{
                    hasData = true;

                    // Add title and metadata rows
                    const metadataRows = [
                        [`PID Annotator - ${{section.name}}`],
                        [`Exported: ${{new Date().toLocaleString()}}`],
                        [] // Blank row
                    ];

                    // Combine metadata, headers, and data
                    const wsData = [
                        ...metadataRows,
                        result.headers,
                        ...result.data
                    ];

                    // Create worksheet
                    const ws = XLSX.utils.aoa_to_sheet(wsData);

                    // Set column widths
                    const colWidths = result.headers.map((header, idx) => {{
                        let maxWidth = header.length;
                        result.data.forEach(row => {{
                            const cellValue = String(row[idx] || '');
                            maxWidth = Math.max(maxWidth, cellValue.length);
                        }});
                        return {{ wch: Math.min(maxWidth + 2, 50) }};
                    }});
                    ws['!cols'] = colWidths;

                    // Merge cells for title
                    ws['!merges'] = [
                        {{ s: {{ r: 0, c: 0 }}, e: {{ r: 0, c: result.headers.length - 1 }} }},
                        {{ s: {{ r: 1, c: 0 }}, e: {{ r: 1, c: result.headers.length - 1 }} }}
                    ];

                    // Add worksheet to workbook
                    XLSX.utils.book_append_sheet(wb, ws, section.name);
                }}
            }});

            if (!hasData) {{
                alert('No visible data to export. Please adjust your filters.');
                return;
            }}

            // Generate and download
            const filename = `pid_annotator_filtered_${{timestamp}}.xlsx`;
            XLSX.writeFile(wb, filename);
        }}

        // --- Unmatched Tags Filter & Export ---

        function getTagPart(tag, partIndex) {{
            // Split tag by - or . delimiters
            const parts = tag.split(/[-.]/);
            if (partIndex >= 0 && partIndex < parts.length) {{
                return parts[partIndex].toUpperCase();
            }}
            return null;
        }}

        function applyUnmatchedFilter() {{
            const partSelect = document.getElementById('unmatchedPartSelect');
            const excludeInput = document.getElementById('unmatchedExcludeInput');
            if (!partSelect || !excludeInput) return;

            const partIndex = parseInt(partSelect.value) - 1; // Convert 1-based to 0-based
            const rawValues = excludeInput.value.trim();

            if (!rawValues) {{
                clearUnmatchedFilter();
                return;
            }}

            const excludeValues = rawValues.split(',')
                .map(v => v.trim().toUpperCase())
                .filter(v => v.length > 0);

            if (excludeValues.length === 0) {{
                clearUnmatchedFilter();
                return;
            }}

            const table = document.getElementById('unmatchedTable');
            if (!table) return;

            const rows = table.querySelectorAll('tbody tr');
            let hiddenCount = 0;
            let visibleCount = 0;

            rows.forEach(row => {{
                const tagCell = row.querySelector('.tag-name');
                if (!tagCell) return;

                const tag = tagCell.textContent.trim();
                const partValue = getTagPart(tag, partIndex);

                if (partValue && excludeValues.includes(partValue)) {{
                    row.style.display = 'none';
                    row.dataset.unmatchedExcluded = 'true';
                    hiddenCount++;
                }} else {{
                    // Only show if not hidden by search filter
                    if (row.dataset.unmatchedExcluded === 'true') {{
                        row.style.display = '';
                        delete row.dataset.unmatchedExcluded;
                    }}
                    visibleCount++;
                }}
            }});

            // Update status
            const status = document.getElementById('unmatchedFilterStatus');
            if (status) {{
                status.textContent = `Excluded ${{hiddenCount}} tags | Showing ${{visibleCount}} tags`;
            }}

            // Update badge
            const badge = document.getElementById('unmatchedBadge');
            if (badge) {{
                badge.textContent = visibleCount;
            }}

            // Update export info
            updateUnmatchedExportInfo(visibleCount);
        }}

        function clearUnmatchedFilter() {{
            const table = document.getElementById('unmatchedTable');
            if (!table) return;

            const excludeInput = document.getElementById('unmatchedExcludeInput');
            if (excludeInput) excludeInput.value = '';

            const rows = table.querySelectorAll('tbody tr');
            let totalCount = 0;

            rows.forEach(row => {{
                row.style.display = '';
                delete row.dataset.unmatchedExcluded;
                totalCount++;
            }});

            const status = document.getElementById('unmatchedFilterStatus');
            if (status) status.textContent = '';

            const badge = document.getElementById('unmatchedBadge');
            if (badge) badge.textContent = totalCount;

            updateUnmatchedExportInfo(totalCount);
        }}

        function updateUnmatchedExportInfo(count) {{
            const info = document.getElementById('unmatchedExportInfo');
            if (info) {{
                info.textContent = count > 0 ? `${{count}} tags will be exported` : '';
            }}
        }}

        function exportUnmatchedToExcel() {{
            if (typeof XLSX === 'undefined') {{
                alert('Excel export library not loaded. Please check your internet connection.');
                return;
            }}

            const table = document.getElementById('unmatchedTable');
            if (!table) {{
                alert('No unmatched tags table found.');
                return;
            }}

            const tbody = table.querySelector('tbody');
            const rows = Array.from(tbody.querySelectorAll('tr'));
            const visibleRows = rows.filter(row => row.style.display !== 'none');

            if (visibleRows.length === 0) {{
                alert('No visible unmatched tags to export. All tags have been excluded.');
                return;
            }}

            const wb = XLSX.utils.book_new();

            // Build data array with headers
            const headers = ['Tag', 'Pages', 'Occurrences'];
            const data = visibleRows.map(row => {{
                return Array.from(row.cells).map(cell => cell.textContent.trim());
            }});

            // Metadata rows
            const metadataRows = [
                ['PID Annotator - Unmatched Tags (In PDF, Not in Excel)'],
                [`Exported: ${{new Date().toLocaleString()}}`],
                []
            ];

            const wsData = [...metadataRows, headers, ...data];
            const ws = XLSX.utils.aoa_to_sheet(wsData);

            // Set column widths
            const colWidths = headers.map((header, idx) => {{
                let maxWidth = header.length;
                data.forEach(row => {{
                    const cellValue = String(row[idx] || '');
                    maxWidth = Math.max(maxWidth, cellValue.length);
                }});
                return {{ wch: Math.min(maxWidth + 2, 50) }};
            }});
            ws['!cols'] = colWidths;

            // Merge title rows
            ws['!merges'] = [
                {{ s: {{ r: 0, c: 0 }}, e: {{ r: 0, c: headers.length - 1 }} }},
                {{ s: {{ r: 1, c: 0 }}, e: {{ r: 1, c: headers.length - 1 }} }}
            ];

            XLSX.utils.book_append_sheet(wb, ws, 'Unmatched Tags');

            // Build filename from PDF name(s)
            const pdfNames = reportData.pdf_filenames || ['unknown'];
            const pdfBaseName = pdfNames.map(name => {{
                return name.replace(new RegExp('[.]pdf$', 'i'), '');
            }}).join('_');
            const filename = `${{pdfBaseName}}_Unmatched_Tags.xlsx`;

            XLSX.writeFile(wb, filename);
        }}

        // Initialize export info on load
        (function() {{
            const table = document.getElementById('unmatchedTable');
            if (table) {{
                const rowCount = table.querySelectorAll('tbody tr').length;
                updateUnmatchedExportInfo(rowCount);
            }}
        }})();
    </script>
</body>
</html>
'''

    return html
