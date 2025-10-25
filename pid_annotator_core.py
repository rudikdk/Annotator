#!/usr/bin/env python3
"""
Core PID annotation functionality without GUI dependencies
- Replaced "stamp" feature with "watermark" implemented via ReportLab + PyPDF2
- Removed background options for the former stamp feature
- Optimized with parallel indexing, memory management, and streaming for large files
"""

import os
import fitz  # PyMuPDF
import pandas as pd
import re
import time
import gc
from collections import defaultdict
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
import xlrd
from xlrd import open_workbook

# Watermark overlay libs
from reportlab.pdfgen import canvas
from PyPDF2 import PdfReader, PdfWriter

# Report generation
from report_template import generate_html_report

# Configuration for optimization features
PARALLEL_INDEXING_ENABLED = True
MAX_WORKERS = max(1, (os.cpu_count() or 2) - 1)  # Leave one core free
STREAMING_THRESHOLD_MB = 50  # Enable streaming for files larger than this
MEMORY_CLEANUP_BATCH_SIZE = 100  # Clean memory after processing this many pages
PROGRESS_UPDATE_INTERVAL = 2  # Minimum percent change to trigger progress update


class TagMatchingConfig:
    """
    Configuration for customizable tag matching behavior.

    Attributes:
        preset: Preset name ('default', 'match_all', 'custom')
        min_parts: Minimum number of tag parts (default: 3)
        max_parts: Maximum number of tag parts (default: 5)
        separators: List of allowed separator characters (default: ['-', '.'])
        min_part_length: Minimum characters per part (default: 1)
        max_part_length: Maximum characters per part (default: 5)
        allow_partial_match: Allow partial tag matching (default: False)
        custom_regex: Optional custom regex pattern (overrides other settings)
    """

    def __init__(self, preset='default', min_parts=3, max_parts=5,
                 separators=None, min_part_length=1, max_part_length=5,
                 allow_partial_match=False, custom_regex=None):
        self.preset = preset
        self.min_parts = min_parts
        self.max_parts = max_parts
        self.separators = separators if separators is not None else ['-', '.']
        self.min_part_length = min_part_length
        self.max_part_length = max_part_length
        self.allow_partial_match = allow_partial_match
        self.custom_regex = custom_regex

    @classmethod
    def from_dict(cls, config_dict):
        """Create TagMatchingConfig from dictionary."""
        if not config_dict:
            return cls()  # Return default config

        return cls(
            preset=config_dict.get('preset', 'default'),
            min_parts=config_dict.get('min_parts', 3),
            max_parts=config_dict.get('max_parts', 5),
            separators=config_dict.get('separators', ['-', '.']),
            min_part_length=config_dict.get('min_part_length', 1),
            max_part_length=config_dict.get('max_part_length', 5),
            allow_partial_match=config_dict.get('allow_partial_match', False),
            custom_regex=config_dict.get('custom_regex', None)
        )

    def to_dict(self):
        """Convert TagMatchingConfig to dictionary."""
        return {
            'preset': self.preset,
            'min_parts': self.min_parts,
            'max_parts': self.max_parts,
            'separators': self.separators,
            'min_part_length': self.min_part_length,
            'max_part_length': self.max_part_length,
            'allow_partial_match': self.allow_partial_match,
            'custom_regex': self.custom_regex
        }

    @classmethod
    def get_default_preset(cls):
        """Get default preset configuration."""
        return cls(preset='default', min_parts=3, max_parts=5,
                   separators=['-', '.'], min_part_length=1, max_part_length=5,
                   allow_partial_match=False, custom_regex=None)

    @classmethod
    def get_match_all_preset(cls):
        """Get match-all preset configuration."""
        return cls(preset='match_all', min_parts=1, max_parts=10,
                   separators=['-', '.', '_', '/', ':'], min_part_length=1, max_part_length=20,
                   allow_partial_match=False, custom_regex=None)


def generate_regex_pattern(config):
    """
    Generate regex pattern from TagMatchingConfig.

    Args:
        config: TagMatchingConfig instance

    Returns:
        compiled regex pattern
    """
    # If custom regex is provided, use it
    if config.custom_regex:
        try:
            return re.compile(config.custom_regex, re.IGNORECASE)
        except re.error as e:
            print(f"Invalid custom regex pattern: {e}. Falling back to default pattern.")
            # Fall back to default pattern
            config = TagMatchingConfig.get_default_preset()

    # Escape separators for regex
    escaped_separators = [re.escape(sep) for sep in config.separators]
    separator_pattern = '|'.join(escaped_separators)

    # Build part pattern
    part_pattern = f'[A-Za-z0-9]{{{config.min_part_length},{config.max_part_length}}}'

    # Build full pattern based on min/max parts
    if config.min_parts == config.max_parts:
        # Exact number of parts
        parts = [part_pattern] * config.min_parts
        pattern = f'\\b{f"(?:{separator_pattern})".join(parts)}\\b'
    else:
        # Variable number of parts
        # Start with minimum parts (required)
        required_parts = [part_pattern] * config.min_parts
        required_pattern = f"(?:{separator_pattern})".join(required_parts)

        # Add optional parts
        optional_count = config.max_parts - config.min_parts
        optional_pattern = f"(?:(?:{separator_pattern}){part_pattern}){{0,{optional_count}}}"

        pattern = f'\\b{required_pattern}{optional_pattern}\\b'

    try:
        return re.compile(pattern, re.IGNORECASE)
    except re.error as e:
        print(f"Error compiling generated regex pattern: {e}. Using default pattern.")
        # Fall back to default pattern
        default_pattern = r'\b[A-Za-z0-9]{1,5}[-\.][A-Za-z0-9]{1,5}[-\.][A-Za-z0-9]{1,5}(?:[-\.][A-Za-z0-9]{1,5}){0,2}\b'
        return re.compile(default_pattern, re.IGNORECASE)


def convert_tag_format(tag, from_delimiter="-", to_delimiter="."):
    """Convert tag from one delimiter format to another."""
    return tag.replace(from_delimiter, to_delimiter)


def parse_tag_format(tag, allowed_delimiters=["-", "."]):
    """
    Parse a tag string and determine its format.

    Args:
        tag: The tag string to parse
        allowed_delimiters: List of allowed delimiter characters

    Returns:
        dict: Contains 'valid' (boolean), 'delimiter' (str), 'parts' (list), and 'count' (int)
    """
    if tag.lower() == "nan" or tag == "":
        return {
            'valid': False,
            'delimiter': None,
            'parts': [],
            'count': 0
        }

    # Determine which delimiter is used
    delimiter = None
    for delim in allowed_delimiters:
        if delim in tag:
            delimiter = delim
            break

    # If no delimiter found, tag is invalid
    if not delimiter:
        return {
            'valid': False,
            'delimiter': None,
            'parts': [tag],
            'count': 1
        }

    # Split the tag into parts
    parts = tag.split(delimiter)

    # Check if all parts are non-empty
    all_parts_valid = all(part.strip() for part in parts)

    return {
        'valid': all_parts_valid and len(parts) >= 2,  # At least 2 parts required
        'delimiter': delimiter,
        'parts': parts,
        'count': len(parts)
    }


def parse_tag_parts(tag, allowed_delimiters=["-", "."]):
    """
    Parse a tag and return its individual parts.
    Supports both '-' and '.' delimiters.

    Args:
        tag: Tag string to parse (e.g., "A-HV-001-A" or "A.HV.001.A")
        allowed_delimiters: List of allowed delimiter characters

    Returns:
        list: List of tag parts (e.g., ["A", "HV", "001", "A"])
    """
    if not tag or tag.lower() == "nan" or tag == "":
        return []

    # Find which delimiter is used
    for delim in allowed_delimiters:
        if delim in tag:
            parts = tag.split(delim)
            # Filter out empty parts
            return [part.strip() for part in parts if part.strip()]

    # No delimiter found, return single-part list
    return [tag.strip()] if tag.strip() else []


def apply_tag_filters(tag, filters, filter_logic="AND"):
    """
    Check if a tag matches the filter criteria.

    Args:
        tag: Tag string to check
        filters: List of filter rules, each containing:
            - part: Part number (1-based index, e.g., 2 for second part)
            - value: Value to match
            - match_type: "exact" or "contains"
            - action: "include" or "exclude"
        filter_logic: "AND" or "OR" - how to combine multiple filters

    Returns:
        bool: True if tag should be processed, False if it should be filtered out
    """
    if not filters:
        return True  # No filters means all tags pass

    if not tag or tag.lower() == "nan":
        return False

    # Parse tag into parts
    tag_parts = parse_tag_parts(tag)

    if not tag_parts:
        return False

    # Separate include and exclude filters
    include_filters = [f for f in filters if f.get('action') == 'include']
    exclude_filters = [f for f in filters if f.get('action') == 'exclude']

    # Check exclude filters first (if any exclude filter matches, tag is excluded)
    for filter_rule in exclude_filters:
        part_index = filter_rule.get('part', 1) - 1  # Convert to 0-based
        value = filter_rule.get('value', '').strip()
        match_type = filter_rule.get('match_type', 'exact')

        # Skip if tag doesn't have this part
        if part_index < 0 or part_index >= len(tag_parts):
            continue

        tag_part = tag_parts[part_index].upper()
        value_upper = value.upper()

        # Check if filter matches
        if match_type == 'exact':
            if tag_part == value_upper:
                return False  # Excluded
        elif match_type == 'contains':
            if value_upper in tag_part:
                return False  # Excluded

    # If no include filters, tag passes (wasn't excluded)
    if not include_filters:
        return True

    # Check include filters based on logic
    matches = []
    for filter_rule in include_filters:
        part_index = filter_rule.get('part', 1) - 1  # Convert to 0-based
        value = filter_rule.get('value', '').strip()
        match_type = filter_rule.get('match_type', 'exact')

        # If tag doesn't have this part, it doesn't match this filter
        if part_index < 0 or part_index >= len(tag_parts):
            matches.append(False)
            continue

        tag_part = tag_parts[part_index].upper()
        value_upper = value.upper()

        # Check if filter matches
        if match_type == 'exact':
            matches.append(tag_part == value_upper)
        elif match_type == 'contains':
            matches.append(value_upper in tag_part)
        else:
            matches.append(False)

    # Apply logic
    if filter_logic == "AND":
        return all(matches) if matches else False
    else:  # OR
        return any(matches) if matches else False


def apply_color_rules(tag, row_data, color_rules, default_color="#FFFF00"):
    """
    Apply color rules to a tag and return the appropriate highlight color.
    Colors are applied based on tag part matching - if tag is found and matches a rule, it gets colored.

    Args:
        tag: Tag string to check (e.g., "A-HV-001-A")
        row_data: Pandas Series containing the Excel row data for this tag (unused, kept for compatibility)
        color_rules: List of color rule dictionaries, each containing:
            - part: Part number (1-based index, e.g., 2 for second part)
            - value: Value to match
            - match_type: "exact" or "contains"
            - color: Hex color code (e.g., "#0000FF")
            - id: Unique identifier for the rule
        default_color: Fallback color if no rules match

    Returns:
        tuple: (color_hex, matched_rule_id, conflicts)
            - color_hex: The color to use (or default_color if no match)
            - matched_rule_id: ID of the matched rule (or None)
            - conflicts: List of rule IDs that also matched (empty if no conflicts)
    """
    if not color_rules:
        return default_color, None, []

    if not tag or tag.lower() == "nan":
        return None, None, []

    # Parse tag into parts
    tag_parts = parse_tag_parts(tag)

    if not tag_parts:
        return None, None, []

    # Track all matching rules (last match wins, but we report conflicts)
    matched_color = None
    matched_rule_id = None
    all_matched_rule_ids = []

    # Iterate through rules (last matching rule wins)
    for rule in color_rules:
        part = rule.get('part', 1)
        value = rule.get('value', '').strip()
        match_type = rule.get('match_type', 'exact')
        color = rule.get('color', '#FFFF00')
        rule_id = rule.get('id', '')

        # Check if tag part matches
        part_index = part - 1  # Convert to 0-based

        # Skip if tag doesn't have this part
        if part_index < 0 or part_index >= len(tag_parts):
            continue

        tag_part = tag_parts[part_index].upper()
        value_upper = value.upper()

        # Check if this rule matches
        matches = False
        if match_type == 'exact':
            matches = (tag_part == value_upper)
        elif match_type == 'contains':
            matches = (value_upper in tag_part)

        if matches:
            # Track this match
            all_matched_rule_ids.append(rule_id)
            # Update matched color (last match wins)
            matched_color = color
            matched_rule_id = rule_id

    # Determine conflicts (if more than one rule matched)
    conflicts = all_matched_rule_ids[:-1] if len(all_matched_rule_ids) > 1 else []

    # If no rules matched, use default color
    if not matched_color:
        return default_color, None, []

    return matched_color, matched_rule_id, conflicts


def analyze_tag_parts(excel_path, tag_column, header_row=6, top_n=20):
    """
    Analyze an Excel file and extract the most common values for each tag part.

    Args:
        excel_path: Path to the Excel file
        tag_column: Column name containing the tags
        header_row: Row number containing headers (1-based)
        top_n: Number of top values to return per part

    Returns:
        dict: Contains 'success', 'parts' (dict with part statistics), and 'message'
    """
    try:
        # Support both .xlsx and .xls files
        is_xls = excel_path.lower().endswith('.xls') and not excel_path.lower().endswith('.xlsx')

        if is_xls:
            df = pd.read_excel(excel_path, header=header_row-1, engine='xlrd')
        else:
            df = pd.read_excel(excel_path, header=header_row-1, engine='openpyxl')

        df = df.dropna(axis=1, how="all")

        # Validate tag column
        if tag_column not in df.columns:
            return {
                'success': False,
                'parts': {},
                'message': f"Tag column '{tag_column}' not found in Excel file"
            }

        # Extract all tags
        tags = df[tag_column].dropna().astype(str).str.strip()

        # Track values for each part position
        part_values = defaultdict(lambda: defaultdict(int))  # {part_num: {value: count}}
        max_parts = 0

        for tag in tags:
            if tag and tag.lower() != 'nan':
                parts = parse_tag_parts(tag)
                max_parts = max(max_parts, len(parts))

                for i, part in enumerate(parts):
                    part_num = i + 1  # 1-based
                    part_values[part_num][part.upper()] += 1

        # Build result with top N values for each part
        result_parts = {}
        for part_num in range(1, max_parts + 1):
            if part_num in part_values:
                # Sort by count (descending) and get top N
                sorted_values = sorted(
                    part_values[part_num].items(),
                    key=lambda x: x[1],
                    reverse=True
                )[:top_n]

                result_parts[f"part{part_num}"] = [
                    {"value": value, "count": count}
                    for value, count in sorted_values
                ]
            else:
                result_parts[f"part{part_num}"] = []

        return {
            'success': True,
            'parts': result_parts,
            'message': f"Analyzed {len(tags)} tags with up to {max_parts} parts"
        }

    except Exception as e:
        return {
            'success': False,
            'parts': {},
            'message': f"Error analyzing tag parts: {str(e)}"
        }


def is_valid_tag(tag, config=None):
    """
    Check if a tag has the correct format with delimiters.

    Args:
        tag: Tag string to validate
        config: Optional TagMatchingConfig instance (uses default if not provided)

    Returns:
        bool: True if tag is valid according to config
    """
    if tag.lower() == "nan" or tag == "":
        return False

    # Use default config if not provided
    if config is None:
        config = TagMatchingConfig.get_default_preset()

    # Use the new parse_tag_format function with config separators
    tag_info = parse_tag_format(tag, allowed_delimiters=config.separators)

    # Check if tag meets the configured requirements
    return tag_info['valid'] and config.min_parts <= tag_info['count'] <= config.max_parts


def _index_page_range(doc_path, page_start, page_end, tag_pattern, config=None):
    """
    Index a range of pages from a PDF document (worker function for parallel processing).

    Args:
        doc_path: Path to the PDF file
        page_start: Starting page number (inclusive)
        page_end: Ending page number (exclusive)
        tag_pattern: Compiled regex pattern for tag matching
        config: TagMatchingConfig instance

    Returns:
        dict: Partial tag index for this page range
    """
    local_index = defaultdict(list)

    # Use default config if not provided
    if config is None:
        config = TagMatchingConfig.get_default_preset()

    # Open document in worker thread
    doc = fitz.open(doc_path)

    try:
        for page_num in range(page_start, page_end):
            if page_num >= len(doc):
                break

            page = doc[page_num]
            text = page.get_text()

            # Find all potential tags on this page
            matches = tag_pattern.finditer(text)

            for match in matches:
                original_tag = match.group()

                # Skip if not a valid tag format
                if not is_valid_tag(original_tag, config):
                    continue
                
                # Get exact coordinates for this tag occurrence
                rects = page.search_for(original_tag, flags=1)
                
                if rects:  # Only add if we can find the coordinates
                    # Normalize tag for lookup
                    normalized_tag = original_tag.upper()
                    
                    # Store both original format and converted formats
                    tag_variants = [
                        normalized_tag,
                        convert_tag_format(normalized_tag, from_delimiter="-", to_delimiter="."),
                        convert_tag_format(normalized_tag, from_delimiter=".", to_delimiter="-")
                    ]
                    
                    # Add all variants to index
                    for variant in set(tag_variants):
                        local_index[variant].append((page_num, rects, original_tag))
            
            # Memory cleanup after each page
            page = None
        
    finally:
        doc.close()
        # Force garbage collection
        gc.collect()
    
    return dict(local_index)


def build_page_bookmark_map(doc):
    """
    Build a mapping of page numbers to their corresponding bookmark titles.

    Args:
        doc: PyMuPDF document object

    Returns:
        dict: Mapping of page numbers (0-based) to bookmark titles
              Format: {page_num: bookmark_title}
    """
    page_bookmark_map = {}

    try:
        # Get table of contents (bookmarks/outline)
        # Returns list of [level, title, page] where page is 1-based
        toc = doc.get_toc()

        if not toc:
            print("No bookmarks found in PDF")
            return page_bookmark_map

        print(f"Found {len(toc)} bookmarks in PDF")

        # Build map: assign each bookmark to its page and all following pages
        # until the next bookmark of same or higher level
        for i, (level, title, page_1based) in enumerate(toc):
            # Convert to 0-based page number
            page_num = page_1based - 1

            # Find the end page for this bookmark
            # (page before next bookmark of same or higher level)
            end_page = len(doc) - 1  # Default to last page

            for j in range(i + 1, len(toc)):
                next_level, _, next_page_1based = toc[j]
                if next_level <= level:
                    end_page = next_page_1based - 2  # Page before next bookmark (0-based)
                    break

            # Assign this bookmark title to all pages in its range
            # But prefer deeper level bookmarks (child over parent)
            for p in range(page_num, min(end_page + 1, len(doc))):
                # Only update if no bookmark assigned yet, or current is deeper level
                if p not in page_bookmark_map:
                    page_bookmark_map[p] = title
                else:
                    # Keep the most specific (deepest) bookmark
                    # We process in order, so later deeper bookmarks will override
                    current_bookmark_idx = next(
                        (idx for idx, (_, ttl, _) in enumerate(toc)
                         if ttl == page_bookmark_map[p]),
                        -1
                    )
                    if current_bookmark_idx >= 0:
                        current_level = toc[current_bookmark_idx][0]
                        if level > current_level:
                            page_bookmark_map[p] = title

        print(f"Mapped bookmarks to {len(page_bookmark_map)} pages")

    except Exception as e:
        print(f"Error building page-bookmark map: {e}")

    return page_bookmark_map


def build_tag_index(doc, update_progress=None, use_parallel=None, pdf_path=None, config=None):
    """
    Build a comprehensive index of all potential tags found in the PDF.
    Supports both sequential and parallel processing modes.

    Args:
        doc: PyMuPDF document object
        update_progress: Optional progress callback function
        use_parallel: Override parallel processing (True/False/None=auto)
        pdf_path: Path to PDF file (required for parallel processing)
        config: TagMatchingConfig instance (uses default if not provided)

    Returns:
        dict: Index mapping normalized tag strings to their locations
              Format: {tag_normalized: [(page_num, rects, original_tag), ...]}
    """
    print("Building comprehensive tag index...")
    start_time = time.time()

    # Use default config if not provided
    if config is None:
        config = TagMatchingConfig.get_default_preset()

    total_pages = len(doc)

    # Determine if we should use parallel processing
    if use_parallel is None:
        use_parallel = PARALLEL_INDEXING_ENABLED and total_pages > 20

    # Generate tag pattern from config
    tag_pattern = generate_regex_pattern(config)
    print(f"Using tag matching pattern: {tag_pattern.pattern}")

    if use_parallel and pdf_path and total_pages > 20:
        print(f"Using parallel indexing with {MAX_WORKERS} workers for {total_pages} pages")
        return _build_tag_index_parallel(pdf_path, total_pages, tag_pattern, update_progress, config)
    else:
        print(f"Using sequential indexing for {total_pages} pages")
        return _build_tag_index_sequential(doc, total_pages, tag_pattern, update_progress, config)


def _build_tag_index_sequential(doc, total_pages, tag_pattern, update_progress=None, config=None):
    """Sequential tag indexing (original implementation)."""
    tag_index = defaultdict(list)
    last_reported_progress = -1

    # Use default config if not provided
    if config is None:
        config = TagMatchingConfig.get_default_preset()

    for page_num in range(total_pages):
        # Update progress with throttling
        if update_progress:
            progress = int((page_num / total_pages) * 100) if total_pages else 100
            if progress - last_reported_progress >= PROGRESS_UPDATE_INTERVAL:
                update_progress(progress, f"Indexing page {page_num + 1}/{total_pages}...")
                last_reported_progress = progress

        page = doc[page_num]
        text = page.get_text()

        # Find all potential tags on this page
        matches = tag_pattern.finditer(text)

        for match in matches:
            original_tag = match.group()

            if not is_valid_tag(original_tag, config):
                continue
            
            rects = page.search_for(original_tag, flags=1)
            
            if rects:
                normalized_tag = original_tag.upper()
                tag_variants = [
                    normalized_tag,
                    convert_tag_format(normalized_tag, from_delimiter="-", to_delimiter="."),
                    convert_tag_format(normalized_tag, from_delimiter=".", to_delimiter="-")
                ]
                
                for variant in set(tag_variants):
                    tag_index[variant].append((page_num, rects, original_tag))
        
        # Memory cleanup every MEMORY_CLEANUP_BATCH_SIZE pages
        if (page_num + 1) % MEMORY_CLEANUP_BATCH_SIZE == 0:
            page = None
            gc.collect()
            if hasattr(fitz, 'TOOLS'):
                try:
                    fitz.TOOLS.store_shrink(100)
                except:
                    pass
    
    total_tags = sum(len(locations) for locations in tag_index.values())
    print(f"Sequential indexing complete. Found {total_tags} tag instances across {len(tag_index)} unique tags.")
    return dict(tag_index)


def _build_tag_index_parallel(pdf_path, total_pages, tag_pattern, update_progress=None, config=None):
    """Parallel tag indexing using ThreadPoolExecutor."""
    tag_index = defaultdict(list)
    index_lock = Lock()
    completed_pages = [0]
    last_reported_progress = [-1]

    # Use default config if not provided
    if config is None:
        config = TagMatchingConfig.get_default_preset()

    # Calculate chunk size (distribute pages across workers)
    chunk_size = max(10, total_pages // (MAX_WORKERS * 2))
    chunks = [(i, min(i + chunk_size, total_pages)) for i in range(0, total_pages, chunk_size)]

    def process_chunk(chunk_info):
        """Process a chunk and update progress."""
        start_page, end_page = chunk_info
        local_result = _index_page_range(pdf_path, start_page, end_page, tag_pattern, config)
        
        # Update global progress
        with index_lock:
            completed_pages[0] += (end_page - start_page)
            if update_progress:
                progress = int((completed_pages[0] / total_pages) * 100)
                if progress - last_reported_progress[0] >= PROGRESS_UPDATE_INTERVAL:
                    update_progress(progress, f"Indexed {completed_pages[0]}/{total_pages} pages (parallel)...")
                    last_reported_progress[0] = progress
        
        return local_result
    
    # Execute parallel processing
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = [executor.submit(process_chunk, chunk) for chunk in chunks]
        
        for future in as_completed(futures):
            try:
                local_index = future.result()
                
                # Merge local results into global index
                with index_lock:
                    for tag_variant, locations in local_index.items():
                        tag_index[tag_variant].extend(locations)
                        
            except Exception as e:
                print(f"Error in parallel indexing chunk: {e}")
    
    total_tags = sum(len(locations) for locations in tag_index.values())
    print(f"Parallel indexing complete. Found {total_tags} tag instances across {len(tag_index)} unique tags.")
    return dict(tag_index)


def _hex_to_rgb01(hex_color):
    """Convert hex like '#RRGGBB' to (r,g,b) floats in 0..1. Returns black on failure."""
    try:
        color_hex = hex_color.lstrip('#')
        if len(color_hex) == 6:
            return (
                int(color_hex[0:2], 16) / 255.0,
                int(color_hex[2:4], 16) / 255.0,
                int(color_hex[4:6], 16) / 255.0
            )
    except Exception:
        pass
    return (0, 0, 0)


def apply_color_rules_to_all_text(doc, tag_index, color_rules, default_highlight_color="#FFFF00",
                                  enable_default_color=True, excel_tags_set=None,
                                  excel_constraint_mode=False, excel_constraint_logic="AND",
                                  progress_callback=None):
    """
    Apply color rules to ALL text in the PDF that matches tag patterns, with optional Excel constraints.

    Args:
        doc: PyMuPDF document object
        tag_index: Pre-built tag index from build_tag_index()
        color_rules: List of color rule dictionaries
        default_highlight_color: Default color for tags that don't match any rule
        enable_default_color: Whether to use default color for unmatched tags
        excel_tags_set: Set of tags from Excel (after filtering)
        excel_constraint_mode: Whether to constrain coloring to Excel tags
        excel_constraint_logic: "AND" or "OR" logic for Excel constraint
        progress_callback: Optional progress callback function

    Returns:
        dict: Statistics about colored tags including page-level breakdown
    """
    if not color_rules:
        return {'total_tags': 0, 'colored_tags': 0, 'page_stats': {}}

    colored_count = 0
    total_count = 0
    page_stats = {}  # {page_num: {'colored_count': X, 'total_count': Y}}

    # Initialize excel_tags_set if not provided
    if excel_tags_set is None:
        excel_tags_set = set()

    # Process all tags found in the PDF
    for tag_text, locations in tag_index.items():
        total_count += len(locations)

        # Check if tag is in Excel list
        in_excel = tag_text in excel_tags_set

        # Determine color for this tag
        highlight_color, matched_rule_id, conflicts = apply_color_rules(
            tag_text, None, color_rules, default_highlight_color if enable_default_color else None
        )

        # Determine if this tag should be colored based on constraint mode
        should_color = False
        if not excel_constraint_mode:
            # No constraint - color if rules match
            should_color = (highlight_color is not None)
        elif excel_constraint_logic == 'AND':
            # AND logic: Tag must be in Excel AND match a color rule
            should_color = in_excel and (matched_rule_id is not None)
        elif excel_constraint_logic == 'OR':
            # OR logic: Color if tag matches rules OR is in Excel
            should_color = (highlight_color is not None) or in_excel
            # If in Excel but no rule matched, use default color
            if in_excel and not highlight_color and enable_default_color:
                highlight_color = default_highlight_color

        # Track page-level statistics for all tags
        for page_num, rects, original_tag in locations:
            # Initialize page stats if not exists
            if page_num not in page_stats:
                page_stats[page_num] = {'colored_count': 0, 'total_count': 0}

            # Increment total count for this page
            page_stats[page_num]['total_count'] += 1

        if not should_color or not highlight_color:
            continue

        # Apply color to all occurrences of this tag
        for page_num, rects, original_tag in locations:
            try:
                page = doc[page_num]

                # Add highlight annotation
                hl = page.add_highlight_annot(rects)

                # Convert hex color to RGB
                try:
                    r, g, b = _hex_to_rgb01(highlight_color)
                    hl.set_colors(stroke=(r, g, b))
                except Exception:
                    hl.set_colors(stroke=(1, 1, 0))  # Fallback to yellow

                hl.update()
                colored_count += 1

                # Increment colored count for this page
                page_stats[page_num]['colored_count'] += 1

            except Exception as e:
                print(f"Error coloring tag {tag_text} on page {page_num}: {e}")
                continue

    return {
        'total_tags': total_count,
        'colored_tags': colored_count,
        'page_stats': page_stats
    }


def process_annotations_from_index(
    doc,
    df,
    tag_index,
    tag_col,
    column_color_pairs=None,
    selected_comment_columns=None,
    annotation_type="highlight_only",
    watermark_enabled=False,
    watermark_attribute="",
    watermark_text_color="#000000",
    max_tags=None,
    update_progress=None,
    watermark_items_by_page=None,
    default_highlight_color="#FFFF00",
    header_row=6,
    config=None,
    tag_filters=None,
    filter_logic="AND",
    page_bookmark_map=None,
    color_rules=None
):
    """
    Process annotations using the pre-built tag index for optimal performance.

    Args:
        doc: PyMuPDF document object
        df: DataFrame with Excel data
        tag_index: Pre-built tag index from build_tag_index()
        tag_col: Column name containing tags
        column_color_pairs: list of (column, color) - DEPRECATED, use color_rules instead
        selected_comment_columns: which columns to include in notes
        annotation_type: "highlight_only" or "note_only"
        watermark_enabled: enable watermark placement
        watermark_attribute: Excel column to render as watermark text
        watermark_text_color: color for watermark text (hex)
        max_tags: optional limit
        update_progress: progress callback
        watermark_items_by_page: dict(page_num -> [ {text,x,y,font_size} ])
        default_highlight_color: hex color used when no color rules match but attribute has value
        header_row: Excel header row number (1-based) for row number tracking
        config: TagMatchingConfig instance (uses default if not provided)
        tag_filters: List of filter rules to apply to tags (optional)
        filter_logic: "AND" or "OR" for combining multiple filters (default: "AND")
        page_bookmark_map: dict mapping page numbers to bookmark titles (optional)
        color_rules: List of color rule dicts (part, value, match_type, color, attribute_column, id)

    Returns:
        tuple: (found_tags, skipped_tags, processed_tags_set, report_data)
    """
    print("Processing annotations using tag index...")
    start_time = time.time()

    # Use default config if not provided
    if config is None:
        config = TagMatchingConfig.get_default_preset()

    if watermark_items_by_page is None:
        watermark_items_by_page = defaultdict(list)

    found_tags = 0
    skipped_tags = 0
    processed_tags = set()

    # Initialize report data structure
    report_data = {
        'found': [],  # List of {tag, pages, occurrence_count, excel_row}
        'not_found': [],  # List of {tag, excel_row, reason}
        'duplicates': {},  # Dict of {tag: [excel_rows]}
        'validation_warnings': [],  # List of {tag, excel_row, warning}
        'color_conflicts': []  # List of {tag, excel_row, matched_rule_id, conflict_rule_ids}
    }

    # Track which tags appear multiple times in Excel (duplicates)
    tag_excel_rows = defaultdict(list)  # {tag: [row_numbers]}

    # Limit processing if max_tags specified
    total_to_process = min(max_tags, len(df)) if max_tags and max_tags > 0 else len(df)

    for i, (df_index, row) in enumerate(df.iterrows()):
        # Calculate Excel row number (df index is 0-based, Excel row is header_row + 1 + df_index)
        excel_row = header_row + 1 + df_index

        if max_tags and i >= max_tags:
            print(f"Reached maximum number of tags ({max_tags}). Stopping processing.")
            break

        if update_progress:
            progress = int((i / total_to_process) * 100) if total_to_process else 100
            update_progress(progress, f"Processing tag {i+1}/{total_to_process}...")

        tag = str(row[tag_col]).strip()

        # Track all tags and their Excel rows (for duplicate detection)
        if tag and tag.lower() != 'nan':
            tag_excel_rows[tag].append(excel_row)

        # Skip empty or invalid tags
        if not tag or not is_valid_tag(tag, config):
            if tag:  # Only log if tag exists but is invalid
                print(f"Skipping invalid tag: {tag}")
                skipped_tags += 1
                # Add validation warning to report
                tag_info = parse_tag_format(tag, config.separators)
                if not tag_info['valid']:
                    report_data['validation_warnings'].append({
                        'tag': tag,
                        'excel_row': excel_row,
                        'warning': f"Invalid tag format - has {tag_info['count']} parts, needs {config.min_parts}-{config.max_parts}"
                    })
            continue

        # Apply tag filters if provided
        if tag_filters:
            if not apply_tag_filters(tag, tag_filters, filter_logic):
                print(f"Skipping tag due to filter: {tag}")
                skipped_tags += 1
                # Skip filtered tags entirely - do NOT add to report
                # Filtered tags should not appear in the report at all
                continue

        # Skip already processed tags (duplicate in Excel)
        if tag in processed_tags:
            print(f"Skipping duplicate tag: {tag}")
            skipped_tags += 1
            # Duplicate will be tracked in tag_excel_rows and processed at the end
            continue

        # Normalize tag for index lookup
        normalized_tag = tag.upper()

        # Try to find tag in index (check multiple variants)
        tag_variants = [
            normalized_tag,
            convert_tag_format(normalized_tag, from_delimiter="-", to_delimiter="."),
            convert_tag_format(normalized_tag, from_delimiter=".", to_delimiter="-")
        ]

        tag_locations = []
        for variant in tag_variants:
            if variant in tag_index:
                tag_locations = tag_index[variant]
                break

        # If partial matching is enabled and no exact match found, try partial matching
        if not tag_locations and config.allow_partial_match:
            for index_tag, locations in tag_index.items():
                # Check if the index tag starts with any of our tag variants
                if any(index_tag.startswith(variant) for variant in tag_variants):
                    tag_locations.extend(locations)
            # Remove duplicates if multiple partial matches found
            if tag_locations:
                seen = set()
                unique_locations = []
                for loc in tag_locations:
                    loc_key = (loc[0], tuple(loc[1]))  # (page_num, tuple of rects)
                    if loc_key not in seen:
                        seen.add(loc_key)
                        unique_locations.append(loc)
                tag_locations = unique_locations
                print(f"Found tag via partial matching: {tag} ({len(tag_locations)} occurrences)")

        if not tag_locations:
            # Tag not found in PDF - add to report
            report_data['not_found'].append({
                'tag': tag,
                'excel_row': excel_row,
                'reason': 'not_in_pdf'
            })
            continue  # Tag not found in PDF
        
        # Create note text
        if selected_comment_columns is not None:
            columns_to_include = [col for col in selected_comment_columns if col in df.columns and col != tag_col]
            if columns_to_include:
                note_text = f"TAG: {tag}\n" + "\n".join(
                    f"{col}: {row[col]}" for col in columns_to_include if pd.notna(row[col])
                )
            else:
                note_text = ""
        else:
            note_text = ""
        
        # Process all occurrences of this tag
        for page_num, rects, original_tag in tag_locations:
            page = doc[page_num]
            
            try:
                # Determine highlight color using new color rules system
                highlight_color = None
                matched_rule_id = None
                conflicts = []

                # Use new color_rules system if provided
                if color_rules:
                    highlight_color, matched_rule_id, conflicts = apply_color_rules(
                        tag, row, color_rules, default_highlight_color
                    )

                    # Track conflicts for reporting
                    if conflicts:
                        report_data['color_conflicts'].append({
                            'tag': tag,
                            'excel_row': excel_row,
                            'matched_rule_id': matched_rule_id,
                            'conflict_rule_ids': conflicts
                        })
                # Fall back to legacy column_color_pairs system for backward compatibility
                elif column_color_pairs:
                    for column, color in column_color_pairs:
                        if column and column in df.columns:
                            if pd.notna(row[column]) and str(row[column]).strip() != "":
                                highlight_color = color
                                break

                # Add highlight annotation if color determined or comments are selected
                should_add_highlight = (highlight_color is not None) or note_text
                if should_add_highlight:
                    hl = page.add_highlight_annot(rects)

                    # Set highlight color
                    if highlight_color:
                        # Convert hex color to RGB
                        try:
                            r, g, b = _hex_to_rgb01(highlight_color)
                            hl.set_colors(stroke=(r, g, b))
                        except Exception:
                            hl.set_colors(stroke=(1, 1, 0))  # Fallback to yellow
                    elif note_text:
                        # No color rule matched, but we have comments - use default
                        try:
                            r, g, b = _hex_to_rgb01(default_highlight_color or "#FFFF00")
                            hl.set_colors(stroke=(r, g, b))
                        except Exception:
                            hl.set_colors(stroke=(1, 1, 0))  # Fallback to yellow

                    # Add comment to highlight
                    if annotation_type == "highlight_only" and note_text:
                        hl.set_info(content=note_text)

                    hl.update()
                
                # Add sticky note if using note_only mode
                if annotation_type == "note_only":
                    r0 = rects[0]
                    note_pos = fitz.Point(r0.x0, r0.y0 - 20)
                    ta = page.add_text_annot(note_pos, note_text, icon="Note")
                    ta.update()
                
                # Collect watermark placements if enabled
                if watermark_enabled and watermark_attribute:
                    # Handle multiple attributes (passed as list or single string)
                    attributes = watermark_attribute if isinstance(watermark_attribute, list) else [watermark_attribute]
                    
                    # Build watermark text from multiple attributes (values only, no attribute names)
                    watermark_parts = []
                    for attr in attributes:
                        if attr and attr in df.columns:
                            attr_value = str(row[attr]).strip()
                            if attr_value and attr_value.lower() != "nan" and attr_value != "":
                                watermark_parts.append(attr_value)  # Only the value, not "attr: value"
                    
                    if watermark_parts:
                        try:
                            r0 = rects[0]

                            # Join all attributes with " / " on a single line
                            # Example: [Val1, Val2, Val3, Val4, Val5] → "Val1 / Val2 / Val3 / Val4 / Val5"
                            wm_text = " / ".join(watermark_parts)
                            
                            # Orientation-aware placement relative to the tag box.
                            # Also record desired text rotation (0 = horizontal, 90 = vertical).
                            font_size = 9
                            char_width = font_size * 0.6
                            text_length = len(wm_text) * char_width  # approximate text width in points
                            margin = 6
                            text_height = font_size
                            
                            r = r0  # first rectangle for the tag
                            is_horizontal = (r.width >= r.height)
                            tag_center = fitz.Point((r.x0 + r.x1) / 2, (r.y0 + r.y1) / 2)
                            
                            if is_horizontal:
                                # dette er vertical tag  i virkeligheden!
                                wm_width = text_length + (margin * 2)
                                wm_x = tag_center.x - (wm_width / 2)
                                wm_y = r.y1 + 3  # slightly below
                                desired_rotation = 270
                            else:
                                # Place to the right of a vertical tag; draw text vertically (bottom-to-top).
                                wm_x = r.x1 - (margin * 3) # a bit to the right
                                # Anchor at the lower end so, after 90° rotation, the text centers on the tag
                                wm_y = tag_center.y - (text_length / 2)
                                desired_rotation = 180
                            
                            watermark_items_by_page[page_num].append({
                                'text': wm_text,
                                'x': wm_x,
                                'y': wm_y,
                                'font_size': font_size,
                                'rotation': desired_rotation
                            })
                        except Exception as e:
                            print(f"Error collecting watermark for tag '{tag}': {e}")
                            continue
                            
            except Exception as e:
                print(f"Error processing tag '{tag}' on page {page_num}: {e}")
                continue
        
        if tag_locations:
            found_tags += 1
            processed_tags.add(tag)
            print(f"Found and processed tag: {tag} ({len(tag_locations)} occurrences)")

            # Add to found tags in report
            # Extract unique pages where this tag was found
            pages = sorted(set(page_num for page_num, _, _ in tag_locations))

            # Extract bookmark titles for each page
            bookmarks = []
            if page_bookmark_map:
                for page_num in pages:
                    bookmark = page_bookmark_map.get(page_num, "N/A")
                    bookmarks.append(bookmark)
            else:
                bookmarks = ["N/A"] * len(pages)

            report_data['found'].append({
                'tag': tag,
                'pages': pages,
                'bookmarks': bookmarks,
                'occurrence_count': len(tag_locations),
                'excel_row': excel_row
            })

    # Process duplicate detection
    for tag, excel_rows in tag_excel_rows.items():
        if len(excel_rows) > 1:
            report_data['duplicates'][tag] = excel_rows

    elapsed_time = time.time() - start_time
    print(f"Annotation processing completed in {elapsed_time:.2f}s")
    print(f"Report: {len(report_data['found'])} found, {len(report_data['not_found'])} not found, {len(report_data['duplicates'])} duplicates, {len(report_data['validation_warnings'])} warnings")

    return found_tags, skipped_tags, processed_tags, report_data


def reload_excel_columns(excel_path, header_row):
    """
    Reload Excel columns with a new header row

    Args:
        excel_path: Path to the Excel file
        header_row: Row number containing headers (1-based)

    Returns:
        dict: Contains 'success', 'columns', 'message', and 'default_tag_column'
    """
    try:
        if header_row < 1:
            return {
                'success': False,
                'columns': [],
                'message': 'Header row must be 1 or greater',
                'default_tag_column': None
            }

        # Support both .xlsx and .xls files
        # .xls files can be read for processing but cannot be annotated (Excel annotation disabled for .xls)
        is_xls = excel_path.lower().endswith('.xls') and not excel_path.lower().endswith('.xlsx')

        if is_xls:
            # Use xlrd engine for .xls files
            df = pd.read_excel(excel_path, header=header_row-1, engine='xlrd')
        else:
            # Use openpyxl engine for .xlsx files
            df = pd.read_excel(excel_path, header=header_row-1, engine='openpyxl')

        df = df.dropna(axis=1, how="all")  # Remove empty columns
        
        columns = list(df.columns)
        
        # Set default tag column to column G (index 6) if it exists, otherwise first column
        default_tag_column = None
        if len(columns) > 6:
            default_tag_column = columns[6]  # Column G
        elif columns:
            default_tag_column = columns[0]  # First available column
        
        message = f"Successfully loaded {len(columns)} columns from header row {header_row}"
        if header_row != 6:
            message += f" (changed from default row 6)"
        
        return {
            'success': True,
            'columns': columns,
            'message': message,
            'default_tag_column': default_tag_column
        }
        
    except Exception as e:
        return {
            'success': False,
            'columns': [],
            'message': f'Error loading Excel columns: {str(e)}',
            'default_tag_column': None
        }


def _get_file_size_mb(filepath):
    """Get file size in megabytes."""
    try:
        return os.path.getsize(filepath) / (1024 * 1024)
    except:
        return 0


def annotate_pdf_with_progress(
    pdf_path,
    excel_path,
    out_path,
    column_color_pairs=None,
    max_tags=None,
    tag_column=None,
    header_row=6,
    selected_comment_columns=None,
    task_id=None,
    progress_callback=None,
    annotation_type="highlight_only",
    watermark_enabled=False,
    watermark_attribute="",
    watermark_text_color="#000000",
    watermark_background_enabled=False,
    default_highlight_color="#FFFF00",
    use_streaming=None,
    tag_matching_config=None,
    tag_filters=None,
    filter_logic="AND",
    color_rules=None,
    enable_default_color=True,
    excel_constraint_mode=False,
    excel_constraint_logic="AND"
):
    """
    Annotate PDF with tags from Excel file with progress tracking.
    Automatically uses streaming mode for large files (>50MB by default).

    Args:
        pdf_path: Path to the PDF file
        excel_path: Path to the Excel file
        out_path: Path for the output annotated PDF
        column_color_pairs: List of (column, color) tuples for conditional highlighting
        max_tags: Maximum number of tags to process (for testing)
        tag_column: Column name containing the tags
        header_row: Row number containing headers (1-based, default is 6)
        selected_comment_columns: List of column names to include in comments (None = all columns)
        task_id: Task ID for progress tracking
        progress_callback: Function to call for progress updates
        annotation_type: Type of annotation to use ("highlight_only" or "note_only")
        watermark_enabled: Whether to enable watermark feature
        watermark_attribute: Column name to use for watermark text
        watermark_text_color: Text color for watermark (hex format)
        use_streaming: Override streaming mode (True/False/None=auto)
    """
    # Determine if we should use streaming mode
    if use_streaming is None:
        file_size_mb = _get_file_size_mb(pdf_path)
        use_streaming = file_size_mb > STREAMING_THRESHOLD_MB
        if use_streaming:
            print(f"[OPTIMIZATION] Large file detected ({file_size_mb:.1f}MB). Using streaming mode.")

    if use_streaming:
        return _annotate_pdf_streaming(
            pdf_path, excel_path, out_path, column_color_pairs, max_tags,
            tag_column, header_row, selected_comment_columns, task_id,
            progress_callback, annotation_type, watermark_enabled,
            watermark_attribute, watermark_text_color, watermark_background_enabled,
            default_highlight_color, tag_matching_config, tag_filters, filter_logic,
            color_rules, enable_default_color, excel_constraint_mode, excel_constraint_logic
        )
    else:
        return _annotate_pdf_standard(
            pdf_path, excel_path, out_path, column_color_pairs, max_tags,
            tag_column, header_row, selected_comment_columns, task_id,
            progress_callback, annotation_type, watermark_enabled,
            watermark_attribute, watermark_text_color, watermark_background_enabled,
            default_highlight_color, tag_matching_config, tag_filters, filter_logic,
            color_rules, enable_default_color, excel_constraint_mode, excel_constraint_logic
        )


def _annotate_pdf_standard(
    pdf_path, excel_path, out_path, column_color_pairs, max_tags,
    tag_column, header_row, selected_comment_columns, task_id,
    progress_callback, annotation_type, watermark_enabled,
    watermark_attribute, watermark_text_color, watermark_background_enabled,
    default_highlight_color, tag_matching_config=None, tag_filters=None, filter_logic="AND",
    color_rules=None, enable_default_color=True, excel_constraint_mode=False, excel_constraint_logic="AND"
):
    """Standard annotation mode - loads entire PDF into memory."""
    # Use default config if not provided
    if tag_matching_config is None:
        config = TagMatchingConfig.get_default_preset()
    else:
        config = TagMatchingConfig.from_dict(tag_matching_config) if isinstance(tag_matching_config, dict) else tag_matching_config
    def update_progress(progress, status):
        print(f"[CORE DEBUG] update_progress called: progress={progress}%, status='{status}'")
        if progress_callback:
            try:
                progress_callback(task_id, progress, status)
                print(f"[CORE DEBUG] progress_callback executed successfully for task {task_id}")
            except Exception as e:
                print(f"[CORE ERROR] progress_callback failed: {e}")
        else:
            print(f"[CORE WARNING] No progress_callback provided")
    
    print(f"\n--- Starting annotation process ---")
    print(f"PDF: {pdf_path}")
    print(f"Excel: {excel_path}")
    print(f"Output: {out_path}")
    print(f"Watermark enabled: {watermark_enabled}")
    if watermark_enabled:
        print(f"Watermark attribute: {watermark_attribute}")
        print(f"Watermark text color: {watermark_text_color}")
    
    # Read Excel with configurable header row
    update_progress(2, "Starting Excel file validation...")
    print(f"Loading Excel file...")
    print(f"Using header row: {header_row}")

    # Support both .xlsx and .xls files for reading
    is_xls = excel_path.lower().endswith('.xls') and not excel_path.lower().endswith('.xlsx')

    update_progress(5, "Reading Excel file...")
    if is_xls:
        # Use xlrd engine for .xls files
        df = pd.read_excel(excel_path, header=header_row-1, engine='xlrd')
    else:
        # Use openpyxl engine for .xlsx files
        df = pd.read_excel(excel_path, header=header_row-1, engine='openpyxl')

    update_progress(8, "Processing Excel data...")
    df = df.dropna(axis=1, how="all")  # Remove empty columns
    print(f"Excel loaded successfully. Found {len(df)} rows.")

    update_progress(10, "Excel file loaded successfully")

    # Use selected tag column or default to column G (index 6)
    if tag_column and tag_column in df.columns:
        tag_col = tag_column
    else:
        tag_col = df.columns[6]  # Default to column G
    
    tags = df[tag_col].dropna().astype(str).str.strip()
    print(f"Found {len(tags)} tags in column '{tag_col}'.")

    update_progress(12, "Validating PDF file...")
    print(f"Opening PDF file...")

    update_progress(15, "Loading PDF document...")
    doc = fitz.open(pdf_path)

    update_progress(18, "Analyzing PDF structure...")
    print(f"PDF opened successfully. Contains {len(doc)} pages.")

    update_progress(20, "PDF file ready for processing")

    update_progress(22, "Preparing annotation parameters...")
    print(f"Found {len(tags)} tags in column '{tag_col}'.")

    update_progress(25, "Initializing processing pipeline...")

    update_progress(28, "Setting up annotation engine...")

    # PHASE 1: Build comprehensive tag index (30-60% progress)
    def index_progress_callback(progress, status):
        # Map index progress (0-100%) to overall progress (30-60%)
        overall_progress = 30 + int(progress * 0.3)
        update_progress(overall_progress, status)

    update_progress(30, "Building comprehensive tag index...")
    tag_index = build_tag_index(doc, index_progress_callback, pdf_path=pdf_path, config=config)

    # Build page-to-bookmark mapping for report
    update_progress(60, "Extracting PDF bookmarks...")
    page_bookmark_map = build_page_bookmark_map(doc)

    # Build Excel tags set for color rule filtering
    excel_tags_set = set()
    if tag_filters:
        update_progress(61, "Applying tag filters to Excel data...")
        for tag in tags:
            if apply_tag_filters(tag, tag_filters, filter_logic):
                excel_tags_set.add(tag)
    else:
        excel_tags_set = set(tags)

    # PHASE 1.5: Apply color rules to ALL matching text in PDF (if color rules provided)
    color_stats = None
    if color_rules:
        update_progress(62, "Applying color rules to all matching text...")
        color_stats = apply_color_rules_to_all_text(
            doc, tag_index, color_rules, default_highlight_color,
            enable_default_color=enable_default_color,
            excel_tags_set=excel_tags_set,
            excel_constraint_mode=excel_constraint_mode,
            excel_constraint_logic=excel_constraint_logic
        )
        print(f"Colored {color_stats['colored_tags']} out of {color_stats['total_tags']} tag occurrences based on color rules.")

    # PHASE 2: Process annotations using index (65-90% progress)
    def annotation_progress_callback(progress, status):
        # Map annotation progress (0-100%) to overall progress (65-90%)
        overall_progress = 65 + int(progress * 0.25)
        update_progress(overall_progress, status)

    # Collect watermark placements per page during processing
    watermark_items_by_page = defaultdict(list)

    update_progress(65, "Processing Excel annotations and comments...")
    found_tags, skipped_tags, processed_tags, report_data = process_annotations_from_index(
        doc=doc,
        df=df,
        tag_index=tag_index,
        tag_col=tag_col,
        column_color_pairs=column_color_pairs,
        selected_comment_columns=selected_comment_columns,
        annotation_type=annotation_type,
        watermark_enabled=watermark_enabled,
        watermark_attribute=watermark_attribute,
        watermark_text_color=watermark_text_color,
        max_tags=max_tags,
        update_progress=annotation_progress_callback,
        watermark_items_by_page=watermark_items_by_page,
        default_highlight_color=default_highlight_color,
        header_row=header_row,
        config=config,
        tag_filters=tag_filters,
        filter_logic=filter_logic,
        page_bookmark_map=page_bookmark_map,
        color_rules=color_rules
    )

    # Add page statistics to report data if color rules were applied
    if color_stats and 'page_stats' in color_stats:
        report_data['page_stats'] = color_stats['page_stats']

    # Save intermediate annotated PDF (highlights/notes only)
    update_progress(95, "Saving annotated PDF (pre-watermark)...")
    print(f"Annotation (highlights/notes) complete.")
    print(f"Found {found_tags} tags out of {len(tags)} valid tags.")
    print(f"Skipped {skipped_tags} invalid or duplicate tags.")
    tmp_out_path = f"{out_path}.prewatermark.pdf"
    print(f"Saving pre-watermark PDF to {tmp_out_path}...")
    doc.save(tmp_out_path, incremental=False, deflate=True)
    doc.close()
    print(f"Pre-watermark PDF saved successfully.")

    # Apply watermark overlay via ReportLab + PyPDF2 if enabled
    def _create_overlay_page(width, height, items, color_hex, page_rotation=0, background_enabled=False):
        """Create a single-page PDF overlay with watermark items at specified positions.
        Applies per-item rotation and compensates for the page's /Rotate value so text
        appears in the intended orientation after viewer rotation.

        Args:
            width: Page width in points
            height: Page height in points
            items: List of watermark items with text, position, rotation
            color_hex: Text color in hex format
            page_rotation: Page rotation value (0, 90, 180, 270)
            background_enabled: If True, draw white rectangle behind text
        """
        r, g, b = _hex_to_rgb01(color_hex or "#000000")
        packet = BytesIO()
        c = canvas.Canvas(packet, pagesize=(width, height))
        c.setFillColorRGB(r, g, b)

        # Normalize rotation to [0,360)
        try:
            page_rotation = int(page_rotation) % 360
        except Exception:
            page_rotation = 0

        for it in items:
            text = it.get('text', '')
            x = float(it.get('x', 0))
            y_fitz = float(it.get('y', 0))
            font_size = int(it.get('font_size', 9)) or 9
            desired_rotation = int(it.get('rotation', 0)) % 360

            # Convert PyMuPDF (origin top-left) to ReportLab (origin bottom-left)
            # Adjust baseline by font size for better visual alignment
            y_rl = float(height) - (y_fitz + font_size)

            # Compensate for page rotation so final appearance matches intent
            final_rotation = (desired_rotation - page_rotation) % 360

            c.saveState()
            c.setFont("Helvetica", font_size)

            # Draw white background rectangle if enabled
            if background_enabled:
                # Estimate text dimensions
                text_width = c.stringWidth(text, "Helvetica", font_size)
                padding = 3  # padding in points

                # Draw white filled rectangle before text
                c.setFillColorRGB(1, 1, 1)  # white
                c.translate(x, y_rl)
                if final_rotation:
                    c.rotate(final_rotation)
                # Draw rectangle at origin with padding
                c.rect(-padding, -padding,
                       text_width + 2*padding,
                       font_size + 2*padding,
                       fill=1, stroke=0)
                # Restore color for text
                c.setFillColorRGB(r, g, b)
            else:
                # No background, just translate and rotate
                c.translate(x, y_rl)
                if final_rotation:
                    c.rotate(final_rotation)

            # Draw text at local origin after transform
            c.drawString(0, 0, text)
            c.restoreState()

        c.showPage()
        c.save()
        packet.seek(0)
        return packet.getvalue()

    if watermark_enabled and any(len(v) > 0 for v in watermark_items_by_page.values()):
        update_progress(97, "Applying watermark overlays...")
        print("Applying watermark overlays using ReportLab + PyPDF2...")
        reader = PdfReader(tmp_out_path)
        writer = PdfWriter()

        num_pages = len(reader.pages)
        for p in range(num_pages):
            base_page = reader.pages[p]
            # Compute page size (points)
            try:
                width = float(base_page.mediabox.right) - float(base_page.mediabox.left)
                height = float(base_page.mediabox.top) - float(base_page.mediabox.bottom)
            except Exception:
                # Fallback in case of unexpected mediabox values
                width = 595.0  # A4 width in points
                height = 842.0  # A4 height in points

            items = watermark_items_by_page.get(p, [])
            if items:
                # Respect page rotation so overlay text appears correctly oriented
                try:
                    page_rotation = int(base_page.get('/Rotate', 0)) % 360
                except Exception:
                    page_rotation = 0
                overlay_bytes = _create_overlay_page(width, height, items, watermark_text_color, page_rotation, watermark_background_enabled)
                overlay_reader = PdfReader(BytesIO(overlay_bytes))
                overlay_page = overlay_reader.pages[0]
                # Merge overlay at (0,0)
                base_page.merge_page(overlay_page)

            writer.add_page(base_page)

        with open(out_path, "wb") as f_out:
            writer.write(f_out)

        print("Watermark overlay applied and final PDF written.")
        try:
            os.remove(tmp_out_path)
        except Exception:
            pass
    else:
        # No watermarks: move prewatermark PDF to final output
        print("No watermark overlays to apply; finalizing output.")
        try:
            # Prefer atomic replace when possible
            os.replace(tmp_out_path, out_path)
        except Exception:
            # Fallback to copy-then-remove
            import shutil
            shutil.copyfile(tmp_out_path, out_path)
            try:
                os.remove(tmp_out_path)
            except Exception:
                pass

    print(f"PDF saved successfully to {out_path}.")

    # Return the set of found tags for Excel annotation and report data
    return processed_tags, report_data


def _annotate_pdf_streaming(
    pdf_path, excel_path, out_path, column_color_pairs, max_tags,
    tag_column, header_row, selected_comment_columns, task_id,
    progress_callback, annotation_type, watermark_enabled,
    watermark_attribute, watermark_text_color, watermark_background_enabled,
    default_highlight_color, tag_matching_config=None, tag_filters=None, filter_logic="AND",
    color_rules=None, enable_default_color=True, excel_constraint_mode=False, excel_constraint_logic="AND"
):
    """Streaming annotation mode - for large PDFs, processes in chunks with memory management."""
    # Use default config if not provided
    if tag_matching_config is None:
        config = TagMatchingConfig.get_default_preset()
    else:
        config = TagMatchingConfig.from_dict(tag_matching_config) if isinstance(tag_matching_config, dict) else tag_matching_config

    def update_progress(progress, status):
        print(f"[CORE DEBUG STREAMING] update_progress called: progress={progress}%, status='{status}'")
        if progress_callback:
            try:
                progress_callback(task_id, progress, status)
            except Exception as e:
                print(f"[CORE ERROR] progress_callback failed: {e}")

    print(f"\n--- Starting STREAMING annotation process ---")
    print(f"PDF: {pdf_path}")
    print(f"Using streaming mode for large file optimization")
    
    # Read Excel data
    update_progress(2, "Loading Excel file...")
    # Support both .xlsx and .xls files for reading
    is_xls = excel_path.lower().endswith('.xls') and not excel_path.lower().endswith('.xlsx')

    if is_xls:
        # Use xlrd engine for .xls files
        df = pd.read_excel(excel_path, header=header_row-1, engine='xlrd')
    else:
        # Use openpyxl engine for .xlsx files
        df = pd.read_excel(excel_path, header=header_row-1, engine='openpyxl')

    df = df.dropna(axis=1, how="all")
    
    if tag_column and tag_column in df.columns:
        tag_col = tag_column
    else:
        tag_col = df.columns[6]
    
    update_progress(10, "Excel loaded. Building tag index with streaming...")
    
    # Build tag index using parallel processing (efficient for large files)
    doc = fitz.open(pdf_path)
    total_pages = len(doc)
    
    def index_progress_callback(progress, status):
        overall_progress = 10 + int(progress * 0.4)
        update_progress(overall_progress, f"Streaming index: {status}")

    tag_index = build_tag_index(doc, index_progress_callback, pdf_path=pdf_path, config=config)

    # Build page-to-bookmark mapping for report
    update_progress(50, "Extracting PDF bookmarks...")
    page_bookmark_map = build_page_bookmark_map(doc)

    # Close and reopen document to clear memory
    doc.close()
    gc.collect()
    if hasattr(fitz, 'TOOLS'):
        try:
            fitz.TOOLS.store_shrink(100)
        except:
            pass

    update_progress(52, "Processing annotations in streaming mode...")

    # Reopen for annotation
    doc = fitz.open(pdf_path)
    watermark_items_by_page = defaultdict(list)

    def annotation_progress_callback(progress, status):
        overall_progress = 52 + int(progress * 0.33)
        update_progress(overall_progress, f"Streaming annotations: {status}")

    found_tags, skipped_tags, processed_tags, report_data = process_annotations_from_index(
        doc=doc,
        df=df,
        tag_index=tag_index,
        tag_col=tag_col,
        column_color_pairs=column_color_pairs,
        selected_comment_columns=selected_comment_columns,
        annotation_type=annotation_type,
        watermark_enabled=watermark_enabled,
        watermark_attribute=watermark_attribute,
        watermark_text_color=watermark_text_color,
        max_tags=max_tags,
        update_progress=annotation_progress_callback,
        watermark_items_by_page=watermark_items_by_page,
        default_highlight_color=default_highlight_color,
        header_row=header_row,
        config=config,
        tag_filters=tag_filters,
        filter_logic=filter_logic,
        page_bookmark_map=page_bookmark_map,
        color_rules=color_rules
    )

    update_progress(85, "Saving annotated PDF (streaming mode)...")
    
    # Save with compression
    tmp_out_path = f"{out_path}.prewatermark.pdf"
    doc.save(tmp_out_path, incremental=False, deflate=True, garbage=4, clean=True)
    doc.close()
    
    # Clear memory before watermark phase
    gc.collect()
    if hasattr(fitz, 'TOOLS'):
        try:
            fitz.TOOLS.store_shrink(100)
        except:
            pass
    
    # Apply watermarks if needed (same as standard mode)
    def _create_overlay_page(width, height, items, color_hex, page_rotation=0, background_enabled=False):
        """Create overlay page with optional background (streaming mode version)."""
        r, g, b = _hex_to_rgb01(color_hex or "#000000")
        packet = BytesIO()
        c = canvas.Canvas(packet, pagesize=(width, height))
        c.setFillColorRGB(r, g, b)

        try:
            page_rotation = int(page_rotation) % 360
        except Exception:
            page_rotation = 0

        for it in items:
            text = it.get('text', '')
            x = float(it.get('x', 0))
            y_fitz = float(it.get('y', 0))
            font_size = int(it.get('font_size', 9)) or 9
            desired_rotation = int(it.get('rotation', 0)) % 360
            y_rl = float(height) - (y_fitz + font_size)
            final_rotation = (desired_rotation - page_rotation) % 360

            c.saveState()
            c.setFont("Helvetica", font_size)

            # Draw white background rectangle if enabled
            if background_enabled:
                # Estimate text dimensions
                text_width = c.stringWidth(text, "Helvetica", font_size)
                padding = 3  # padding in points

                # Draw white filled rectangle before text
                c.setFillColorRGB(1, 1, 1)  # white
                c.translate(x, y_rl)
                if final_rotation:
                    c.rotate(final_rotation)
                # Draw rectangle at origin with padding
                c.rect(-padding, -padding,
                       text_width + 2*padding,
                       font_size + 2*padding,
                       fill=1, stroke=0)
                # Restore color for text
                c.setFillColorRGB(r, g, b)
            else:
                # No background, just translate and rotate
                c.translate(x, y_rl)
                if final_rotation:
                    c.rotate(final_rotation)

            # Draw text at local origin after transform
            c.drawString(0, 0, text)
            c.restoreState()

        c.showPage()
        c.save()
        packet.seek(0)
        return packet.getvalue()
    
    if watermark_enabled and any(len(v) > 0 for v in watermark_items_by_page.values()):
        update_progress(90, "Applying watermarks (streaming mode)...")
        reader = PdfReader(tmp_out_path)
        writer = PdfWriter()
        
        num_pages = len(reader.pages)
        for p in range(num_pages):
            base_page = reader.pages[p]
            
            try:
                width = float(base_page.mediabox.right) - float(base_page.mediabox.left)
                height = float(base_page.mediabox.top) - float(base_page.mediabox.bottom)
            except Exception:
                width = 595.0
                height = 842.0
            
            items = watermark_items_by_page.get(p, [])
            if items:
                try:
                    page_rotation = int(base_page.get('/Rotate', 0)) % 360
                except Exception:
                    page_rotation = 0
                
                overlay_bytes = _create_overlay_page(width, height, items, watermark_text_color, page_rotation, watermark_background_enabled)
                overlay_reader = PdfReader(BytesIO(overlay_bytes))
                overlay_page = overlay_reader.pages[0]
                base_page.merge_page(overlay_page)
            
            writer.add_page(base_page)
            
            # Memory cleanup every 50 pages in streaming mode
            if (p + 1) % 50 == 0:
                gc.collect()
        
        with open(out_path, "wb") as f_out:
            writer.write(f_out)
        
        try:
            os.remove(tmp_out_path)
        except Exception:
            pass
    else:
        try:
            os.replace(tmp_out_path, out_path)
        except Exception:
            import shutil
            shutil.copyfile(tmp_out_path, out_path)
            try:
                os.remove(tmp_out_path)
            except Exception:
                pass
    
    update_progress(98, "Streaming annotation complete")
    print(f"Streaming mode: PDF saved successfully to {out_path}.")
    print(f"Found {found_tags} tags, skipped {skipped_tags}")

    return processed_tags, report_data


def convert_xls_to_xlsx_preserving_format(xls_path, xlsx_path):
    """
    Convert .xls file to .xlsx format while preserving all formatting, formulas, and data.

    Args:
        xls_path: Path to the input .xls file
        xlsx_path: Path for the output .xlsx file

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        print(f"[XLS CONVERSION] Converting {xls_path} to {xlsx_path}")

        # Open the .xls file with xlrd
        xls_book = open_workbook(xls_path, formatting_info=True)

        # Create new .xlsx workbook
        xlsx_book = Workbook()
        # Remove the default sheet
        if 'Sheet' in xlsx_book.sheetnames:
            del xlsx_book['Sheet']

        # Process each sheet
        for sheet_idx in range(xls_book.nsheets):
            xls_sheet = xls_book.sheet_by_index(sheet_idx)
            xlsx_sheet = xlsx_book.create_sheet(title=xls_sheet.name)

            print(f"[XLS CONVERSION] Processing sheet '{xls_sheet.name}' ({xls_sheet.nrows} rows x {xls_sheet.ncols} cols)")

            # Copy all cells with their values and formatting
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    xls_cell = xls_sheet.cell(row_idx, col_idx)
                    xlsx_cell = xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1)

                    # Copy cell value based on type
                    if xls_cell.ctype == xlrd.XL_CELL_EMPTY:
                        xlsx_cell.value = None
                    elif xls_cell.ctype == xlrd.XL_CELL_TEXT:
                        xlsx_cell.value = xls_cell.value
                    elif xls_cell.ctype == xlrd.XL_CELL_NUMBER:
                        xlsx_cell.value = xls_cell.value
                    elif xls_cell.ctype == xlrd.XL_CELL_DATE:
                        # Convert Excel date to datetime
                        try:
                            date_tuple = xlrd.xldate_as_tuple(xls_cell.value, xls_book.datemode)
                            from datetime import datetime
                            xlsx_cell.value = datetime(*date_tuple)
                        except:
                            xlsx_cell.value = xls_cell.value
                    elif xls_cell.ctype == xlrd.XL_CELL_BOOLEAN:
                        xlsx_cell.value = bool(xls_cell.value)
                    elif xls_cell.ctype == xlrd.XL_CELL_ERROR:
                        xlsx_cell.value = xls_cell.value
                    else:
                        xlsx_cell.value = xls_cell.value

                    # Copy cell formatting
                    try:
                        xls_format = xls_book.format_map.get(xls_cell.xf_index)
                        if xls_format:
                            # Get the XF (extended format) record
                            xf = xls_book.format_map[xls_cell.xf_index]

                            # Copy font formatting
                            try:
                                font_idx = xf.font_index
                                xls_font = xls_book.font_list[font_idx]

                                xlsx_cell.font = Font(
                                    name=xls_font.name,
                                    size=xls_font.height / 20.0,  # Convert to points
                                    bold=xls_font.bold,
                                    italic=xls_font.italic,
                                    underline='single' if xls_font.underline_type else 'none',
                                    color=None  # Color mapping is complex, skip for now
                                )
                            except:
                                pass

                            # Copy alignment
                            try:
                                xlsx_cell.alignment = Alignment(
                                    horizontal=['general', 'left', 'center', 'right', 'fill', 'justify', 'centerContinuous', 'distributed'][xf.alignment.hor_align] if xf.alignment.hor_align < 8 else 'general',
                                    vertical=['top', 'center', 'bottom', 'justify', 'distributed'][xf.alignment.vert_align] if xf.alignment.vert_align < 5 else 'bottom',
                                    wrap_text=bool(xf.alignment.text_wrapped)
                                )
                            except:
                                pass
                    except:
                        pass  # Skip formatting if not available

            # Copy column widths
            try:
                for col_idx in range(xls_sheet.ncols):
                    try:
                        # Get column width from .xls (in 1/256th of character width)
                        if hasattr(xls_sheet, 'colinfo_map') and col_idx in xls_sheet.colinfo_map:
                            col_info = xls_sheet.colinfo_map[col_idx]
                            width_in_chars = col_info.width / 256.0
                            xlsx_sheet.column_dimensions[get_column_letter(col_idx + 1)].width = width_in_chars
                    except:
                        pass
            except:
                pass

            # Copy row heights
            try:
                for row_idx in range(xls_sheet.nrows):
                    try:
                        if hasattr(xls_sheet, 'rowinfo_map') and row_idx in xls_sheet.rowinfo_map:
                            row_info = xls_sheet.rowinfo_map[row_idx]
                            if row_info.height:
                                xlsx_sheet.row_dimensions[row_idx + 1].height = row_info.height / 20.0  # Convert to points
                    except:
                        pass
            except:
                pass

            # Copy merged cells
            try:
                for merged_range in xls_sheet.merged_cells:
                    # merged_range is (rlo, rhi, clo, chi) - row_low, row_high, col_low, col_high
                    rlo, rhi, clo, chi = merged_range
                    # Convert to openpyxl format (1-based)
                    start_cell = f"{get_column_letter(clo + 1)}{rlo + 1}"
                    end_cell = f"{get_column_letter(chi)}{rhi}"
                    xlsx_sheet.merge_cells(f"{start_cell}:{end_cell}")
            except:
                pass

        # Save the .xlsx file
        xlsx_book.save(xlsx_path)
        print(f"[XLS CONVERSION] Successfully converted to {xlsx_path}")

        # Verify the file was created
        if os.path.exists(xlsx_path):
            file_size = os.path.getsize(xlsx_path)
            print(f"[XLS CONVERSION] Output file size: {file_size} bytes")
            return True
        else:
            print(f"[XLS CONVERSION ERROR] Output file was not created")
            return False

    except Exception as e:
        print(f"[XLS CONVERSION ERROR] Failed to convert .xls to .xlsx: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def annotate_excel_with_found_tags(excel_path, out_path, found_tags_set, tag_column, header_row=6, report_data=None):
    """
    Annotate Excel file by highlighting rows based on tag status (duplicates, not found, found)

    Args:
        excel_path: Path to the original Excel file
        out_path: Path for the output annotated Excel file
        found_tags_set: Set of tag strings that were found in the PDF
        tag_column: Column name containing the tags
        header_row: Row number containing headers (1-based, default is 6)
        report_data: Optional dict with 'duplicates', 'not_found', 'found' lists for advanced colorization

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        print(f"\n[EXCEL ANNOTATION] Starting Excel file annotation...")
        print(f"[EXCEL ANNOTATION] Input path: {excel_path}")
        print(f"[EXCEL ANNOTATION] Output path: {out_path}")
        print(f"[EXCEL ANNOTATION] Tag column: {tag_column}")
        print(f"[EXCEL ANNOTATION] Header row: {header_row}")
        print(f"[EXCEL ANNOTATION] Found tags count: {len(found_tags_set)}")
        print(f"[EXCEL ANNOTATION] Report data provided: {report_data is not None}")

        # Verify input file exists
        if not os.path.exists(excel_path):
            print(f"[EXCEL ANNOTATION ERROR] Input file does not exist: {excel_path}")
            return False

        # Handle .xls files by converting to .xlsx first (preserving all formats and data)
        working_path = excel_path
        is_xls_format = excel_path.lower().endswith('.xls') and not excel_path.lower().endswith('.xlsx')

        if is_xls_format:
            print(f"[EXCEL ANNOTATION] Detected .xls format - converting to .xlsx with format preservation...")
            try:
                # Create temporary .xlsx file
                temp_xlsx_path = excel_path + '.temp.xlsx'

                # Use the new conversion function that preserves all formatting
                if not convert_xls_to_xlsx_preserving_format(excel_path, temp_xlsx_path):
                    print(f"[EXCEL ANNOTATION ERROR] Failed to convert .xls to .xlsx")
                    return False

                working_path = temp_xlsx_path
                print(f"[EXCEL ANNOTATION] Successfully converted .xls to temporary .xlsx file: {working_path}")
            except Exception as conv_error:
                print(f"[EXCEL ANNOTATION ERROR] Failed to convert .xls to .xlsx: {str(conv_error)}")
                import traceback
                traceback.print_exc()
                return False

        print(f"[EXCEL ANNOTATION] Loading Excel workbook with openpyxl...")
        # Load Excel file with openpyxl
        workbook = load_workbook(working_path)
        worksheet = workbook.active
        print(f"[EXCEL ANNOTATION] Workbook loaded successfully. Active sheet: {worksheet.title}")

        # Load data with pandas to get the tag column index
        print(f"[EXCEL ANNOTATION] Loading Excel data with pandas...")
        # Always use openpyxl since we've converted .xls to .xlsx if needed
        df = pd.read_excel(working_path, header=header_row-1, engine='openpyxl')
        df = df.dropna(axis=1, how="all")  # Remove empty columns
        print(f"[EXCEL ANNOTATION] Loaded {len(df)} rows and {len(df.columns)} columns")
        print(f"[EXCEL ANNOTATION] Column names: {list(df.columns)}")

        # Validate that tag_column exists in DataFrame
        if tag_column not in df.columns:
            # Try default column G (index 6)
            if len(df.columns) > 6:
                tag_column = df.columns[6]
                print(f"[EXCEL ANNOTATION] Tag column not found, using default column G: {tag_column}")
            else:
                print(f"[EXCEL ANNOTATION ERROR] Tag column '{tag_column}' not found in Excel file")
                print(f"[EXCEL ANNOTATION ERROR] Available columns: {list(df.columns)}")
                return False

        # Define color fills for different statuses
        # Red for duplicates (FF6B6B - light red)
        duplicate_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        # Orange for not found (FFB347 - light orange)
        not_found_fill = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
        # Green for found (90EE90 - light green)
        found_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        # Build lookup sets from report_data if provided
        duplicate_tags = set()
        not_found_tags = set()

        if report_data:
            # Extract duplicate tags from report_data
            if 'duplicates' in report_data:
                duplicate_tags = set(report_data['duplicates'].keys())
                print(f"[EXCEL ANNOTATION] Found {len(duplicate_tags)} duplicate tags in report")

            # Extract not found tags from report_data
            if 'not_found' in report_data:
                not_found_tags = set(item['tag'] for item in report_data['not_found'])
                print(f"[EXCEL ANNOTATION] Found {len(not_found_tags)} not-found tags in report")
        else:
            print(f"[EXCEL ANNOTATION] No report_data provided - using only found_tags_set for colorization")

        print(f"[EXCEL ANNOTATION] Colorization sources:")
        print(f"[EXCEL ANNOTATION]   - Duplicate tags: {len(duplicate_tags)}")
        print(f"[EXCEL ANNOTATION]   - Not found tags: {len(not_found_tags)}")
        print(f"[EXCEL ANNOTATION]   - Found tags: {len(found_tags_set)}")

        # Track statistics
        duplicate_count = 0
        not_found_count = 0
        found_count = 0

        print(f"[EXCEL ANNOTATION] Starting row iteration for colorization...")
        # Iterate through rows and highlight based on status
        for row_idx, row in df.iterrows():
            # Convert 0-based index to 1-based Excel row
            excel_row = row_idx + header_row + 1  # +1 for header row, +1 because Excel is 1-based

            # Use column name directly instead of index for reliable access
            tag_value = str(row[tag_column]).strip()

            # Skip invalid tags
            if not tag_value or tag_value.lower() == 'nan':
                continue

            # Determine the fill color based on tag status
            fill_color = None

            # Priority: Duplicates > Not Found > Found
            if tag_value in duplicate_tags:
                fill_color = duplicate_fill
                duplicate_count += 1
                print(f"Highlighted row {excel_row} as DUPLICATE (tag: {tag_value})")
            elif tag_value in not_found_tags:
                fill_color = not_found_fill
                not_found_count += 1
                print(f"Highlighted row {excel_row} as NOT FOUND (tag: {tag_value})")
            elif tag_value in found_tags_set:
                fill_color = found_fill
                found_count += 1
                print(f"Highlighted row {excel_row} as FOUND (tag: {tag_value})")

            # Apply the fill color to the entire row
            if fill_color:
                for col_idx in range(1, len(df.columns) + 1):  # Excel columns are 1-based
                    cell = worksheet.cell(row=excel_row, column=col_idx)
                    cell.fill = fill_color

        # Save the annotated Excel file
        print(f"[EXCEL ANNOTATION] Saving annotated Excel file to: {out_path}")
        try:
            workbook.save(out_path)
            print(f"[EXCEL ANNOTATION] File saved successfully!")
        except Exception as save_error:
            print(f"[EXCEL ANNOTATION ERROR] Failed to save file: {str(save_error)}")
            import traceback
            traceback.print_exc()
            return False

        # Verify file was created
        if os.path.exists(out_path):
            file_size = os.path.getsize(out_path)
            print(f"[EXCEL ANNOTATION] File verification SUCCESS:")
            print(f"[EXCEL ANNOTATION]   - File exists: {out_path}")
            print(f"[EXCEL ANNOTATION]   - File size: {file_size} bytes")
        else:
            print(f"[EXCEL ANNOTATION ERROR] File does not exist after save operation: {out_path}")
            return False

        print(f"[EXCEL ANNOTATION] Excel annotation complete:")
        print(f"[EXCEL ANNOTATION]   - {duplicate_count} duplicate tags (red - FF6B6B)")
        print(f"[EXCEL ANNOTATION]   - {not_found_count} not found tags (orange - FFB347)")
        print(f"[EXCEL ANNOTATION]   - {found_count} found tags (green - 90EE90)")
        print(f"[EXCEL ANNOTATION]   - Total: {duplicate_count + not_found_count + found_count} rows highlighted")

        # Clean up temporary .xlsx file if it was created from .xls
        if is_xls_format and working_path != excel_path:
            try:
                os.remove(working_path)
                print(f"[EXCEL ANNOTATION] Cleaned up temporary file: {working_path}")
            except Exception as cleanup_error:
                print(f"[EXCEL ANNOTATION WARNING] Could not remove temporary file: {cleanup_error}")

        return True

    except Exception as e:
        print(f"[EXCEL ANNOTATION ERROR] Exception during Excel annotation: {str(e)}")
        import traceback
        traceback.print_exc()

        # Clean up temporary file if it exists
        try:
            if 'is_xls_format' in locals() and is_xls_format and 'working_path' in locals() and working_path != excel_path:
                if os.path.exists(working_path):
                    os.remove(working_path)
                    print(f"[EXCEL ANNOTATION] Cleaned up temporary file after error: {working_path}")
        except Exception:
            pass

        return False


def annotate_pdf(
    pdf_path,
    excel_path,
    out_path,
    column_color_pairs=None,
    max_tags=None,
    tag_column=None,
    header_row=6,
    selected_comment_columns=None,
    annotation_type="highlight_only",
    watermark_enabled=False,
    watermark_attribute="",
    watermark_text_color="#000000",
    default_highlight_color="#FFFF00"
):
    """
    Annotate PDF with tags from Excel file (original function for backward compatibility)
    """
    return annotate_pdf_with_progress(
        pdf_path,
        excel_path,
        out_path,
        column_color_pairs,
        max_tags,
        tag_column,
        header_row,
        selected_comment_columns,
        None,
        None,
        annotation_type,
        watermark_enabled,
        watermark_attribute,
        watermark_text_color,
        default_highlight_color
    )


def annotate_pdf_page_for_preview(
    pdf_path,
    excel_path,
    page_num,
    tag_column=None,
    header_row=6,
    selected_comment_columns=None,
    color_rules=None,
    default_highlight_color="#FFFF00",
    enable_default_color=True,
    excel_constraint_mode=False,
    excel_constraint_logic="AND",
    watermark_enabled=False,
    watermark_attributes=None,
    watermark_text_color="#000000",
    watermark_background_enabled=False,
    tag_matching_config=None,
    tag_filters=None,
    filter_logic="AND",
    annotation_type="highlight_only"
):
    """
    Process ONLY the specified page with full annotation pipeline for preview.
    This function replicates the exact same processing logic as annotate_pdf_with_progress
    but optimized for single-page preview.

    Args:
        pdf_path: Path to the PDF file
        excel_path: Path to the Excel file
        page_num: 0-indexed page number to process
        tag_column: Column name containing the tags
        header_row: Row number containing headers (1-based)
        selected_comment_columns: List of column names to include in comments
        color_rules: List of color rule dictionaries
        default_highlight_color: Default color for unmatched tags
        enable_default_color: Whether to use default color
        excel_constraint_mode: Whether to constrain coloring to Excel tags
        excel_constraint_logic: "AND" or "OR" logic
        watermark_enabled: Whether to enable watermarks
        watermark_attributes: List of column names for watermark text
        watermark_text_color: Hex color for watermark text
        watermark_background_enabled: Whether to add white background to watermarks
        tag_matching_config: TagMatchingConfig instance or dict
        tag_filters: List of filter rules
        filter_logic: "AND" or "OR" for combining filters
        annotation_type: "highlight_only" or "note_only"

    Returns:
        dict: {
            'success': bool,
            'page': PyMuPDF page object with annotations,
            'stats': {
                'total_tags': int,
                'colored_tags': int,
                'comments_added': int,
                'watermarks_added': int
            },
            'message': str
        }
    """
    try:
        print(f"\n[PREVIEW] Starting single-page preview for page {page_num + 1}")

        # Use default config if not provided
        if tag_matching_config is None:
            config = TagMatchingConfig.get_default_preset()
        else:
            config = TagMatchingConfig.from_dict(tag_matching_config) if isinstance(tag_matching_config, dict) else tag_matching_config

        # Generate tag pattern
        tag_pattern = generate_regex_pattern(config)

        # Load Excel data
        is_xls = excel_path.lower().endswith('.xls') and not excel_path.lower().endswith('.xlsx')
        if is_xls:
            df = pd.read_excel(excel_path, header=header_row-1, engine='xlrd')
        else:
            df = pd.read_excel(excel_path, header=header_row-1, engine='openpyxl')

        df = df.dropna(axis=1, how="all")

        # Use selected tag column or default
        if tag_column and tag_column in df.columns:
            tag_col = tag_column
        else:
            tag_col = df.columns[6] if len(df.columns) > 6 else df.columns[0]

        tags = df[tag_col].dropna().astype(str).str.strip()

        # Build Excel tags set with filters
        excel_tags_set = set()
        if tag_filters:
            for tag in tags:
                if apply_tag_filters(tag, tag_filters, filter_logic):
                    excel_tags_set.add(tag)
        else:
            excel_tags_set = set(tags)

        # Open PDF
        doc = fitz.open(pdf_path)

        # Validate page number
        if page_num < 0 or page_num >= len(doc):
            doc.close()
            return {
                'success': False,
                'message': f'Invalid page number: {page_num + 1}. PDF has {len(doc)} pages.'
            }

        page = doc[page_num]

        # Build tag index for ONLY this page
        print(f"[PREVIEW] Building tag index for page {page_num + 1}...")
        page_tag_index = defaultdict(list)

        text = page.get_text()
        matches = tag_pattern.finditer(text)

        for match in matches:
            original_tag = match.group()

            if not is_valid_tag(original_tag, config):
                continue

            rects = page.search_for(original_tag, flags=1)

            if rects:
                normalized_tag = original_tag.upper()
                tag_variants = [
                    normalized_tag,
                    convert_tag_format(normalized_tag, from_delimiter="-", to_delimiter="."),
                    convert_tag_format(normalized_tag, from_delimiter=".", to_delimiter="-")
                ]

                for variant in set(tag_variants):
                    page_tag_index[variant].append((page_num, rects, original_tag))

        tag_index = dict(page_tag_index)
        print(f"[PREVIEW] Found {len(tag_index)} unique tags on page {page_num + 1}")

        # PHASE 1: Apply color rules to ALL matching text on this page
        color_stats = {'total_tags': 0, 'colored_tags': 0}
        if color_rules:
            print(f"[PREVIEW] Applying color rules...")
            for tag_text, locations in tag_index.items():
                color_stats['total_tags'] += len(locations)

                # Check if tag is in Excel list
                in_excel = tag_text in excel_tags_set

                # Determine color for this tag
                highlight_color, matched_rule_id, conflicts = apply_color_rules(
                    tag_text, None, color_rules, default_highlight_color if enable_default_color else None
                )

                # Determine if this tag should be colored based on constraint mode
                should_color = False
                if not excel_constraint_mode:
                    should_color = (highlight_color is not None)
                elif excel_constraint_logic == 'AND':
                    should_color = in_excel and (matched_rule_id is not None)
                elif excel_constraint_logic == 'OR':
                    should_color = (highlight_color is not None) or in_excel
                    if in_excel and not highlight_color and enable_default_color:
                        highlight_color = default_highlight_color

                if not should_color or not highlight_color:
                    continue

                # Apply color to all occurrences
                for pg_num, rects, original_tag in locations:
                    try:
                        hl = page.add_highlight_annot(rects)
                        r, g, b = _hex_to_rgb01(highlight_color)
                        hl.set_colors(stroke=(r, g, b))
                        hl.update()
                        color_stats['colored_tags'] += 1
                    except Exception as e:
                        print(f"[PREVIEW] Error coloring tag {tag_text}: {e}")

        # PHASE 2: Add comment annotations for Excel tags
        comments_added = 0
        watermarks_added = 0
        watermark_items = []

        if selected_comment_columns or watermark_enabled:
            print(f"[PREVIEW] Processing Excel rows for annotations...")

            for idx, row in df.iterrows():
                tag = str(row[tag_col]).strip()
                if not tag or tag.lower() == 'nan':
                    continue

                # Check filters
                if tag_filters and not apply_tag_filters(tag, tag_filters, filter_logic):
                    continue

                # Find this tag in the index
                tag_normalized = tag.upper()
                tag_variants = [
                    tag_normalized,
                    convert_tag_format(tag_normalized, from_delimiter="-", to_delimiter="."),
                    convert_tag_format(tag_normalized, from_delimiter=".", to_delimiter="-")
                ]

                locations_for_tag = []
                for variant in tag_variants:
                    if variant in tag_index:
                        locations_for_tag.extend(tag_index[variant])

                if not locations_for_tag:
                    continue

                # Process each location
                for pg_num, rects, original_tag in locations_for_tag:
                    if pg_num != page_num:
                        continue

                    # Add comment annotation
                    if selected_comment_columns and annotation_type != "note_only":
                        comment_parts = [f"Tag: {tag}"]
                        for col in selected_comment_columns:
                            if col in df.columns:
                                val = row[col]
                                if pd.notna(val):
                                    comment_parts.append(f"{col}: {val}")

                        comment_text = "\n".join(comment_parts)

                        try:
                            annot = page.add_text_annot(
                                rects[0].tl,
                                comment_text,
                                icon="Note"
                            )
                            comments_added += 1
                        except Exception as e:
                            print(f"[PREVIEW] Error adding comment for {tag}: {e}")

                    # Collect watermark data
                    if watermark_enabled and watermark_attributes:
                        watermark_parts = []
                        for attr in watermark_attributes:
                            if attr in df.columns:
                                val = row[attr]
                                if pd.notna(val):
                                    watermark_parts.append(str(val))

                        if watermark_parts:
                            watermark_text = " | ".join(watermark_parts)
                            first_rect = rects[0]
                            watermark_items.append({
                                'text': watermark_text,
                                'x': first_rect.x0,
                                'y': first_rect.y0 - 15,
                                'font_size': 8
                            })

        # PHASE 3: Apply watermarks if enabled
        if watermark_enabled and watermark_items:
            print(f"[PREVIEW] Applying {len(watermark_items)} watermarks...")
            try:
                for item in watermark_items:
                    # Add white background if enabled
                    if watermark_background_enabled:
                        text_width = len(item['text']) * item['font_size'] * 0.6
                        bg_rect = fitz.Rect(
                            item['x'], item['y'] - item['font_size'],
                            item['x'] + text_width, item['y'] + 2
                        )
                        page.draw_rect(bg_rect, color=(1, 1, 1), fill=(1, 1, 1))

                    # Add text
                    r, g, b = _hex_to_rgb01(watermark_text_color)
                    page.insert_text(
                        (item['x'], item['y']),
                        item['text'],
                        fontsize=item['font_size'],
                        color=(r, g, b)
                    )
                    watermarks_added += 1
            except Exception as e:
                print(f"[PREVIEW] Error applying watermarks: {e}")

        print(f"[PREVIEW] Preview complete: {color_stats['colored_tags']} colored, {comments_added} comments, {watermarks_added} watermarks")

        # Commit all changes by saving and reloading the page
        # This ensures watermarks (added with insert_text) are visible in the rendered image
        print(f"[PREVIEW] Committing changes to page...")
        try:
            import io
            buffer = io.BytesIO()
            doc.save(buffer)
            buffer.seek(0)

            # Store page number before closing
            target_page = page_num
            doc.close()

            # Reopen from buffer
            doc = fitz.open(stream=buffer, filetype="pdf")
            page = doc[target_page]

            print(f"[PREVIEW] Page reloaded successfully - all annotations committed")
        except Exception as e:
            print(f"[PREVIEW] Warning: Failed to reload page, watermarks may not be visible: {e}")
            # Continue anyway with original page

        return {
            'success': True,
            'page': page,
            'doc': doc,  # Return doc so caller can close it
            'stats': {
                'total_tags': color_stats['total_tags'],
                'colored_tags': color_stats['colored_tags'],
                'comments_added': comments_added,
                'watermarks_added': watermarks_added
            },
            'message': 'Preview generated successfully'
        }

    except Exception as e:
        import traceback
        traceback.print_exc()
        return {
            'success': False,
            'message': f'Error generating preview: {str(e)}'
        }
