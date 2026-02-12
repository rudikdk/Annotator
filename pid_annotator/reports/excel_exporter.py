"""
Excel Export Module for PID Annotator
Exports filtered tag lists to Excel format with proper formatting and headers

Usage Example:
    from pid_annotator.reports.excel_exporter import export_filtered_list_to_excel
    from flask import send_file

    @app.route('/export_found_tags')
    def export_found():
        found_tags = session.get('found_tags', [])
        excel_buffer = export_filtered_list_to_excel(found_tags, 'found')
        return send_file(
            excel_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='found_tags.xlsx'
        )
"""

from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def export_filtered_list_to_excel(data_list, list_type='found', include_metadata=True):
    """
    Export filtered tag list to Excel format with proper formatting and headers.

    Args:
        data_list: List of tag data (found, not_found, duplicates, etc.)
        list_type: Type of list ('found', 'not_found', 'duplicates', 'validation_warnings')
        include_metadata: Whether to include metadata header rows (default: True)

    Returns:
        BytesIO: Excel file in memory as bytes
    """
    wb = Workbook()
    ws = wb.active

    # Set worksheet title based on list type
    title_map = {
        'found': 'Found Tags',
        'not_found': 'Not Found Tags',
        'duplicates': 'Duplicate Tags',
        'validation_warnings': 'Validation Warnings'
    }
    ws.title = title_map.get(list_type, 'Tag List')

    # Define styles
    header_fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    metadata_fill = PatternFill(start_color='f3f4f6', end_color='f3f4f6', fill_type='solid')
    metadata_font = Font(bold=True, size=10)
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row_idx = 1

    # Add metadata if requested
    if include_metadata:
        ws.merge_cells(f'A{row_idx}:D{row_idx}')
        metadata_cell = ws[f'A{row_idx}']
        metadata_cell.value = f'PID Annotator - {title_map.get(list_type, "Export")}'
        metadata_cell.font = Font(bold=True, size=14, color='2563eb')
        metadata_cell.alignment = Alignment(horizontal='center', vertical='center')
        row_idx += 1

        ws.merge_cells(f'A{row_idx}:D{row_idx}')
        date_cell = ws[f'A{row_idx}']
        date_cell.value = f'Exported: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        date_cell.font = Font(size=10, color='6b7280')
        date_cell.alignment = Alignment(horizontal='center')
        row_idx += 2  # Add blank row

    # Define headers based on list type
    if list_type == 'found':
        headers = ['Tag', 'Pages', 'Bookmarks', 'Occurrences', 'Excel Row']
    elif list_type == 'not_found':
        headers = ['Tag', 'Excel Row', 'Reason']
    elif list_type == 'duplicates':
        headers = ['Tag', 'Excel Rows', 'Count']
    elif list_type == 'validation_warnings':
        headers = ['Tag', 'Excel Row', 'Warning']
    else:
        headers = ['Tag', 'Data']

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border_thin

    header_row = row_idx
    row_idx += 1

    # Write data rows
    for item in data_list:
        if list_type == 'found':
            # Get bookmarks and format them
            bookmarks = item.get('bookmarks', [])
            bookmarks_str = ', '.join(str(b) for b in bookmarks) if bookmarks else 'N/A'

            row_data = [
                item.get('tag', ''),
                ', '.join(map(str, item.get('pages', []))),
                bookmarks_str,
                item.get('occurrence_count', 0),
                item.get('excel_row', '')
            ]
        elif list_type == 'not_found':
            row_data = [
                item.get('tag', ''),
                item.get('excel_row', ''),
                item.get('reason', 'not_in_pdf')
            ]
        elif list_type == 'duplicates':
            row_data = [
                item.get('tag', ''),
                ', '.join(map(str, item.get('excel_rows', []))),
                item.get('count', 0)
            ]
        elif list_type == 'validation_warnings':
            row_data = [
                item.get('tag', ''),
                item.get('excel_row', ''),
                item.get('warning', '')
            ]
        else:
            row_data = [str(item)]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border_thin
            cell.alignment = Alignment(vertical='center')

            # Special formatting for tag column
            if col_idx == 1:
                cell.font = Font(name='Courier New', bold=True, color='2563eb')

        row_idx += 1

    # Auto-adjust column widths
    for col_idx, header in enumerate(headers, 1):
        max_length = len(header)
        column_letter = ws.cell(row=header_row, column=col_idx).column_letter

        # Check data rows for maximum width
        for row in ws.iter_rows(min_row=header_row + 1, max_row=row_idx - 1, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

        # Set column width with padding (max 50 characters)
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Save to BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer
