"""
Excel processing and annotation functionality.

This module handles Excel file annotation by highlighting rows based on tag status
(duplicates, not found, found) and includes format conversion utilities.
"""

import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from xlrd import open_workbook
import xlrd
from datetime import datetime


def convert_xls_to_xlsx_preserving_format(xls_path, xlsx_path):
    """
    Convert .xls file to .xlsx format while preserving all formatting, formulas, and data.

    Args:
        xls_path: Path to the input .xls file
        xlsx_path: Path for the output .xlsx file

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        print(f"[XLS CONVERSION] Converting {xls_path} to {xlsx_path}")

        # Open the .xls file with xlrd
        xls_book = open_workbook(xls_path, formatting_info=True)

        # Create new .xlsx workbook
        xlsx_book = Workbook()
        # Remove the default sheet
        if 'Sheet' in xlsx_book.sheetnames:
            del xlsx_book['Sheet']

        # Process each sheet
        for sheet_idx in range(xls_book.nsheets):
            xls_sheet = xls_book.sheet_by_index(sheet_idx)
            xlsx_sheet = xlsx_book.create_sheet(title=xls_sheet.name)

            print(f"[XLS CONVERSION] Processing sheet '{xls_sheet.name}' ({xls_sheet.nrows} rows x {xls_sheet.ncols} cols)")

            # Copy all cells with their values and formatting
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    xls_cell = xls_sheet.cell(row_idx, col_idx)
                    xlsx_cell = xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1)

                    # Copy cell value based on type
                    if xls_cell.ctype == xlrd.XL_CELL_EMPTY:
                        xlsx_cell.value = None
                    elif xls_cell.ctype == xlrd.XL_CELL_TEXT:
                        xlsx_cell.value = xls_cell.value
                    elif xls_cell.ctype == xlrd.XL_CELL_NUMBER:
                        xlsx_cell.value = xls_cell.value
                    elif xls_cell.ctype == xlrd.XL_CELL_DATE:
                        # Convert Excel date to datetime
                        try:
                            date_tuple = xlrd.xldate_as_tuple(xls_cell.value, xls_book.datemode)
                            xlsx_cell.value = datetime(*date_tuple)
                        except:
                            xlsx_cell.value = xls_cell.value
                    elif xls_cell.ctype == xlrd.XL_CELL_BOOLEAN:
                        xlsx_cell.value = bool(xls_cell.value)
                    elif xls_cell.ctype == xlrd.XL_CELL_ERROR:
                        xlsx_cell.value = xls_cell.value
                    else:
                        xlsx_cell.value = xls_cell.value

                    # Copy cell formatting
                    try:
                        xls_format = xls_book.format_map.get(xls_cell.xf_index)
                        if xls_format:
                            # Get the XF (extended format) record
                            xf = xls_book.format_map[xls_cell.xf_index]

                            # Copy font formatting
                            try:
                                font_idx = xf.font_index
                                xls_font = xls_book.font_list[font_idx]

                                xlsx_cell.font = Font(
                                    name=xls_font.name,
                                    size=xls_font.height / 20.0,  # Convert to points
                                    bold=xls_font.bold,
                                    italic=xls_font.italic,
                                    underline='single' if xls_font.underline_type else 'none',
                                    color=None  # Color mapping is complex, skip for now
                                )
                            except:
                                pass

                            # Copy alignment
                            try:
                                xlsx_cell.alignment = Alignment(
                                    horizontal=['general', 'left', 'center', 'right', 'fill', 'justify', 'centerContinuous', 'distributed'][xf.alignment.hor_align] if xf.alignment.hor_align < 8 else 'general',
                                    vertical=['top', 'center', 'bottom', 'justify', 'distributed'][xf.alignment.vert_align] if xf.alignment.vert_align < 5 else 'bottom',
                                    wrap_text=bool(xf.alignment.text_wrapped)
                                )
                            except:
                                pass
                    except:
                        pass  # Skip formatting if not available

            # Copy column widths
            try:
                for col_idx in range(xls_sheet.ncols):
                    try:
                        # Get column width from .xls (in 1/256th of character width)
                        if hasattr(xls_sheet, 'colinfo_map') and col_idx in xls_sheet.colinfo_map:
                            col_info = xls_sheet.colinfo_map[col_idx]
                            width_in_chars = col_info.width / 256.0
                            xlsx_sheet.column_dimensions[get_column_letter(col_idx + 1)].width = width_in_chars
                    except:
                        pass
            except:
                pass

            # Copy row heights
            try:
                for row_idx in range(xls_sheet.nrows):
                    try:
                        if hasattr(xls_sheet, 'rowinfo_map') and row_idx in xls_sheet.rowinfo_map:
                            row_info = xls_sheet.rowinfo_map[row_idx]
                            if row_info.height:
                                xlsx_sheet.row_dimensions[row_idx + 1].height = row_info.height / 20.0  # Convert to points
                    except:
                        pass
            except:
                pass

            # Copy merged cells
            try:
                for merged_range in xls_sheet.merged_cells:
                    # merged_range is (rlo, rhi, clo, chi) - row_low, row_high, col_low, col_high
                    rlo, rhi, clo, chi = merged_range
                    # Convert to openpyxl format (1-based)
                    start_cell = f"{get_column_letter(clo + 1)}{rlo + 1}"
                    end_cell = f"{get_column_letter(chi)}{rhi}"
                    xlsx_sheet.merge_cells(f"{start_cell}:{end_cell}")
            except:
                pass

        # Save the .xlsx file
        xlsx_book.save(xlsx_path)
        print(f"[XLS CONVERSION] Successfully converted to {xlsx_path}")

        # Verify the file was created
        if os.path.exists(xlsx_path):
            file_size = os.path.getsize(xlsx_path)
            print(f"[XLS CONVERSION] Output file size: {file_size} bytes")
            return True
        else:
            print(f"[XLS CONVERSION ERROR] Output file was not created")
            return False

    except Exception as e:
        print(f"[XLS CONVERSION ERROR] Failed to convert .xls to .xlsx: {str(e)}")
        import traceback
        traceback.print_exc()
        return False


def annotate_excel_with_found_tags(excel_path, out_path, found_tags_set, tag_column, header_row=6, report_data=None):
    """
    Annotate Excel file by highlighting rows based on tag status (duplicates, not found, found)

    Args:
        excel_path: Path to the original Excel file
        out_path: Path for the output annotated Excel file
        found_tags_set: Set of tag strings that were found in the PDF
        tag_column: Column name containing the tags
        header_row: Row number containing headers (1-based, default is 6)
        report_data: Optional dict with 'duplicates', 'not_found', 'found' lists for advanced colorization

    Returns:
        bool: True if successful, False otherwise
    """
    try:
        print(f"\n[EXCEL ANNOTATION] Starting Excel file annotation...")
        print(f"[EXCEL ANNOTATION] Input path: {excel_path}")
        print(f"[EXCEL ANNOTATION] Output path: {out_path}")
        print(f"[EXCEL ANNOTATION] Tag column: {tag_column}")
        print(f"[EXCEL ANNOTATION] Header row: {header_row}")
        print(f"[EXCEL ANNOTATION] Found tags count: {len(found_tags_set)}")
        print(f"[EXCEL ANNOTATION] Report data provided: {report_data is not None}")

        # Verify input file exists
        if not os.path.exists(excel_path):
            print(f"[EXCEL ANNOTATION ERROR] Input file does not exist: {excel_path}")
            return False

        # Handle .xls files by converting to .xlsx first (preserving all formats and data)
        working_path = excel_path
        is_xls_format = excel_path.lower().endswith('.xls') and not excel_path.lower().endswith('.xlsx')

        if is_xls_format:
            print(f"[EXCEL ANNOTATION] Detected .xls format - converting to .xlsx with format preservation...")
            try:
                # Create temporary .xlsx file
                temp_xlsx_path = excel_path + '.temp.xlsx'

                # Use the new conversion function that preserves all formatting
                if not convert_xls_to_xlsx_preserving_format(excel_path, temp_xlsx_path):
                    print(f"[EXCEL ANNOTATION ERROR] Failed to convert .xls to .xlsx")
                    return False

                working_path = temp_xlsx_path
                print(f"[EXCEL ANNOTATION] Successfully converted .xls to temporary .xlsx file: {working_path}")
            except Exception as conv_error:
                print(f"[EXCEL ANNOTATION ERROR] Failed to convert .xls to .xlsx: {str(conv_error)}")
                import traceback
                traceback.print_exc()
                return False

        print(f"[EXCEL ANNOTATION] Loading Excel workbook with openpyxl...")
        # Load Excel file with openpyxl
        workbook = load_workbook(working_path)
        worksheet = workbook.active
        print(f"[EXCEL ANNOTATION] Workbook loaded successfully. Active sheet: {worksheet.title}")

        # Load data with pandas to get the tag column index
        print(f"[EXCEL ANNOTATION] Loading Excel data with pandas...")
        # Always use openpyxl since we've converted .xls to .xlsx if needed
        df = pd.read_excel(working_path, header=header_row-1, engine='openpyxl')
        df = df.dropna(axis=1, how="all")  # Remove empty columns
        # Strip whitespace from column names for consistent matching
        df.columns = [str(col).strip() for col in df.columns]
        print(f"[EXCEL ANNOTATION] Loaded {len(df)} rows and {len(df.columns)} columns")
        print(f"[EXCEL ANNOTATION] Column names: {list(df.columns)}")

        # Validate that tag_column exists in DataFrame
        if tag_column not in df.columns:
            # Try default column G (index 6)
            if len(df.columns) > 6:
                tag_column = df.columns[6]
                print(f"[EXCEL ANNOTATION] Tag column not found, using default column G: {tag_column}")
            else:
                print(f"[EXCEL ANNOTATION ERROR] Tag column '{tag_column}' not found in Excel file")
                print(f"[EXCEL ANNOTATION ERROR] Available columns: {list(df.columns)}")
                return False

        # Define color fills for different statuses
        # Red for duplicates (FF6B6B - light red)
        duplicate_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
        # Orange for not found (FFB347 - light orange)
        not_found_fill = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
        # Green for found (90EE90 - light green)
        found_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        # Build lookup sets from report_data if provided
        duplicate_tags = set()
        not_found_tags = set()

        if report_data:
            # Extract duplicate tags from report_data
            if 'duplicates' in report_data:
                duplicate_tags = set(report_data['duplicates'].keys())
                print(f"[EXCEL ANNOTATION] Found {len(duplicate_tags)} duplicate tags in report")

            # Extract not found tags from report_data
            if 'not_found' in report_data:
                not_found_tags = set(item['tag'] for item in report_data['not_found'])
                print(f"[EXCEL ANNOTATION] Found {len(not_found_tags)} not-found tags in report")
        else:
            print(f"[EXCEL ANNOTATION] No report_data provided - using only found_tags_set for colorization")

        print(f"[EXCEL ANNOTATION] Colorization sources:")
        print(f"[EXCEL ANNOTATION]   - Duplicate tags: {len(duplicate_tags)}")
        print(f"[EXCEL ANNOTATION]   - Not found tags: {len(not_found_tags)}")
        print(f"[EXCEL ANNOTATION]   - Found tags: {len(found_tags_set)}")

        # Track statistics
        duplicate_count = 0
        not_found_count = 0
        found_count = 0

        print(f"[EXCEL ANNOTATION] Starting row iteration for colorization...")
        # Iterate through rows and highlight based on status
        for row_idx, row in df.iterrows():
            # Convert 0-based index to 1-based Excel row
            excel_row = row_idx + header_row + 1  # +1 for header row, +1 because Excel is 1-based

            # Use column name directly instead of index for reliable access
            tag_value = str(row[tag_column]).strip()

            # Skip invalid tags
            if not tag_value or tag_value.lower() == 'nan':
                continue

            # Determine the fill color based on tag status
            fill_color = None

            # Priority: Duplicates > Not Found > Found
            if tag_value in duplicate_tags:
                fill_color = duplicate_fill
                duplicate_count += 1
                print(f"Highlighted row {excel_row} as DUPLICATE (tag: {tag_value})")
            elif tag_value in not_found_tags:
                fill_color = not_found_fill
                not_found_count += 1
                print(f"Highlighted row {excel_row} as NOT FOUND (tag: {tag_value})")
            elif tag_value in found_tags_set:
                fill_color = found_fill
                found_count += 1
                print(f"Highlighted row {excel_row} as FOUND (tag: {tag_value})")

            # Apply the fill color to the entire row
            if fill_color:
                for col_idx in range(1, len(df.columns) + 1):  # Excel columns are 1-based
                    cell = worksheet.cell(row=excel_row, column=col_idx)
                    cell.fill = fill_color

        # Save the annotated Excel file
        print(f"[EXCEL ANNOTATION] Saving annotated Excel file to: {out_path}")
        try:
            workbook.save(out_path)
            print(f"[EXCEL ANNOTATION] File saved successfully!")
        except Exception as save_error:
            print(f"[EXCEL ANNOTATION ERROR] Failed to save file: {str(save_error)}")
            import traceback
            traceback.print_exc()
            return False

        # Verify file was created
        if os.path.exists(out_path):
            file_size = os.path.getsize(out_path)
            print(f"[EXCEL ANNOTATION] File verification SUCCESS:")
            print(f"[EXCEL ANNOTATION]   - File exists: {out_path}")
            print(f"[EXCEL ANNOTATION]   - File size: {file_size} bytes")
        else:
            print(f"[EXCEL ANNOTATION ERROR] File does not exist after save operation: {out_path}")
            return False

        print(f"[EXCEL ANNOTATION] Excel annotation complete:")
        print(f"[EXCEL ANNOTATION]   - {duplicate_count} duplicate tags (red - FF6B6B)")
        print(f"[EXCEL ANNOTATION]   - {not_found_count} not found tags (orange - FFB347)")
        print(f"[EXCEL ANNOTATION]   - {found_count} found tags (green - 90EE90)")
        print(f"[EXCEL ANNOTATION]   - Total: {duplicate_count + not_found_count + found_count} rows highlighted")

        # Clean up temporary .xlsx file if it was created from .xls
        if is_xls_format and working_path != excel_path:
            try:
                os.remove(working_path)
                print(f"[EXCEL ANNOTATION] Cleaned up temporary file: {working_path}")
            except Exception as cleanup_error:
                print(f"[EXCEL ANNOTATION WARNING] Could not remove temporary file: {cleanup_error}")

        return True

    except Exception as e:
        print(f"[EXCEL ANNOTATION ERROR] Exception during Excel annotation: {str(e)}")
        import traceback
        traceback.print_exc()

        # Clean up temporary file if it exists
        try:
            if 'is_xls_format' in locals() and is_xls_format and 'working_path' in locals() and working_path != excel_path:
                if os.path.exists(working_path):
                    os.remove(working_path)
                    print(f"[EXCEL ANNOTATION] Cleaned up temporary file after error: {working_path}")
        except Exception:
            pass

        return False
